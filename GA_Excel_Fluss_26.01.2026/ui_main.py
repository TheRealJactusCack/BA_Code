# FILE: ui_main.py
# UI-Einstiegspunkt: lädt Excel-Layout, konfiguriert GA-Parameter, startet GA und zeigt Population/Bestlösung.

from __future__ import annotations

import pyqtgraph
from PyQt6.QtCore import (Qt, pyqtSlot)
from PyQt6.QtGui import QFont

# Qt Widgets für Fenster, Layouts, Controls, Dialoge
from PyQt6.QtWidgets import (QWidget,QVBoxLayout,QHBoxLayout, QGroupBox, QLabel,
    QPushButton, QSpinBox, QDoubleSpinBox, QMessageBox, QApplication,
    QFileDialog, QComboBox, QGridLayout,
)

import copy
import config
from helpers import update_grid_counts
from ga_engine import run_ga
from ui_canvas import PopulationCanvas, BestDialog
from excel_import import apply_excel_layout_to_config

class MainWindow(QWidget):
    # Initialisiert das Hauptfenster und baut die komplette UI (Excel-Controls, GA-Controls, Canvas, Status)
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Fabrikplaner GA")
        self.setMinimumSize(1024, 768)

        # Merker: ob ein Layout bereits erfolgreich geladen wurde + welcher Excel-Pfad aktiv ist
        self.layout_loaded = False
        self.xlsx_path: str | None = None

        überschrift_font = QFont()
        überschrift_font.setPointSize(14)
        überschrift_font.setFamily("Arial")
        überschrift_font.setBold(True)

        root = QVBoxLayout(self)
        controls = QVBoxLayout()
        root.addLayout(controls)

        # Block: Excel Layout laden/neu laden + Sheet auswählen
        layout_group = QGroupBox("Excel")

        controls.addWidget(layout_group)
        layout_v = QVBoxLayout(layout_group)
        Ebutton_layout = QHBoxLayout()
        Sheet_layout = QHBoxLayout()

        self.excel_file_label = QLabel("Keine Excel Datei geladen")
        self.sheet_combo = QComboBox()
        self.sheet_combo.setFixedSize(100, 25)
        self.sheet_combo.setEnabled(False)
        self.sheet_combo.currentTextChanged.connect(self._on_sheet_changed)

        self.load_excel_btn = QPushButton("Excel laden")
        self.load_excel_btn.clicked.connect(self.load_excel)
        Ebutton_layout.addWidget(self.load_excel_btn)

        self.reload_excel_btn = QPushButton("Layout neu laden")
        self.reload_excel_btn.clicked.connect(self.reload_excel)
        self.reload_excel_btn.setEnabled(False)
        Ebutton_layout.addWidget(self.reload_excel_btn)

        self.machine_info_label = QLabel(f"Maschinen {config.MACHINE_COUNT}")
        Sheet_label = QLabel("Sheet auswählen:")
        Sheet_label.setFixedSize(110, 25)
        Sheet_layout.addWidget(Sheet_label)
        Sheet_layout.addWidget(self.sheet_combo)
        Sheet_layout.addWidget(self.excel_file_label)

        layout_v.addLayout(Ebutton_layout)
        layout_v.addLayout(Sheet_layout)

        layout_v.addWidget(self.machine_info_label)

        # Block: GA-Parameter (Population, Elite, Mutation) + Start/Stop UI
        ga_group = QGroupBox("Parameter anpassen und Starten")
        controls.addWidget(ga_group)
        ga_v = QHBoxLayout(ga_group)

        pop_group = QGroupBox()
        pop_group.setFixedSize(100, 75)

        elite_group = QGroupBox()
        elite_group.setFixedSize(100, 75)

        mut_group = QGroupBox()
        mut_group.setFixedSize(100, 75)

        std_group = QGroupBox()
        std_group.setFixedSize(100, 75)

        rot_group = QGroupBox()
        rot_group.setFixedSize(100, 75)

        swap_group = QGroupBox()
        swap_group.setFixedSize(100, 75)

        gen_group = QGroupBox()
        gen_group.setFixedSize(100, 75)
        
        self.population_size_spin = QSpinBox()
        self.population_size_spin.setRange(1, 1000)
        self.population_size_spin.setValue(config.POPULATION_SIZE)
        pop_group.setLayout(QVBoxLayout())
        pop_group.layout().addWidget(QLabel("Population"),alignment=Qt.AlignmentFlag.AlignHCenter)
        pop_group.layout().addWidget(self.population_size_spin)
        ga_v.addWidget(pop_group)

        self.elite_keep_spin = QSpinBox()
        self.elite_keep_spin.setRange(1, 1000)
        self.elite_keep_spin.setValue(config.ELITE_KEEP)
        elite_group.setLayout(QVBoxLayout())
        elite_group.layout().addWidget(QLabel("Eliten behalten"),alignment=Qt.AlignmentFlag.AlignHCenter)
        elite_group.layout().addWidget(self.elite_keep_spin)
        ga_v.addWidget(elite_group)

        self.mutation_prob_spin = QDoubleSpinBox()
        self.mutation_prob_spin.setRange(0.0, 1.0)
        self.mutation_prob_spin.setValue(config.MUTATION_PROB)
        self.mutation_prob_spin.setSingleStep(0.01)
        mut_group.setLayout(QVBoxLayout())
        mut_group.layout().addWidget(QLabel("Mutation Wkt."),alignment=Qt.AlignmentFlag.AlignHCenter)
        mut_group.layout().addWidget(self.mutation_prob_spin)
        ga_v.addWidget(mut_group)

        self.mutation_pos_std_spin = QDoubleSpinBox()
        self.mutation_pos_std_spin.setRange(0.1, 1000.0)
        self.mutation_pos_std_spin.setValue(config.MUTATION_POS_STD)
        self.mutation_pos_std_spin.setSingleStep(0.1)
        std_group.setLayout(QVBoxLayout())
        std_group.layout().addWidget(QLabel("Std. Abweichung"),alignment=Qt.AlignmentFlag.AlignHCenter)
        std_group.layout().addWidget(self.mutation_pos_std_spin)
        ga_v.addWidget(std_group)

        self.mutation_rot_prob_spin = QDoubleSpinBox()
        self.mutation_rot_prob_spin.setRange(0.0, 1.0)
        self.mutation_rot_prob_spin.setValue(config.MUTATION_ROT_PROB)
        self.mutation_rot_prob_spin.setSingleStep(0.01)
        rot_group.setLayout(QVBoxLayout())
        rot_group.layout().addWidget(QLabel("Rotation Wkt."),alignment=Qt.AlignmentFlag.AlignHCenter)
        rot_group.layout().addWidget(self.mutation_rot_prob_spin)
        ga_v.addWidget(rot_group)

        self.mutation_swap_prob_spin = QDoubleSpinBox()
        self.mutation_swap_prob_spin.setRange(0.0, 1.0)
        self.mutation_swap_prob_spin.setValue(config.SWAP_PROB)
        self.mutation_swap_prob_spin.setSingleStep(0.01)
        swap_group.setLayout(QVBoxLayout())
        swap_group.layout().addWidget(QLabel("Swap Wkt."),alignment=Qt.AlignmentFlag.AlignHCenter)
        swap_group.layout().addWidget(self.mutation_swap_prob_spin)
        ga_v.addWidget(swap_group)

        self.generations_spin = QSpinBox()
        self.generations_spin.setRange(1, 10000)
        self.generations_spin.setValue(config.GENERATIONS)
        gen_group.setLayout(QVBoxLayout())
        gen_group.layout().addWidget(QLabel("Generationen"),alignment=Qt.AlignmentFlag.AlignHCenter)
        gen_group.layout().addWidget(self.generations_spin)        
        ga_v.addWidget(gen_group)

        button_group = QVBoxLayout()
        oben_group = QHBoxLayout()
        unten_group = QHBoxLayout()

        # Button: GA starten (blockiert UI während Lauf)
        self.start_btn = QPushButton("GA starten")
        self.start_btn.setFixedSize(100, 35)
        self.start_btn.clicked.connect(self.start_ga)
        oben_group.addWidget(self.start_btn)

        # Button: Stop-Request an GA (setzt config.STOP_REQUESTED)
        self.stop_btn = QPushButton("Abbrechen")
        self.stop_btn.setFixedSize(100, 35)
        self.stop_btn.setEnabled(False)
        oben_group.addWidget(self.stop_btn)

        # Label: Generation Counter + Total
        self.gencounter_label = QLabel(f"0 / {config.GENERATIONS}")
        self.gencounter_label.setFixedSize(100, 35)
        font = QFont()
        font.setPointSize(11)
        font.setFamily("Arial")
        font.setBold(True)
        self.gencounter_label.setFont(font)
        self.gencounter_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        unten_group.addWidget(self.gencounter_label)

        # Letztes GA-Ergebnis für erneutes Anzeigen im Hauptfenster
        self._last_best_ind = None
        self._last_best_score = None

        # Button: Bestes Ergebnis erneut anzeigen (nach GA-Lauf aktiv)
        self.view_best_btn = QPushButton("Bestes Ergebnis")
        self.view_best_btn.setFixedSize(100, 35)
        self.view_best_btn.setEnabled(False)
        self.view_best_btn.clicked.connect(self.view_best_result)
        unten_group.addWidget(self.view_best_btn)

        button_group.addLayout(oben_group)
        button_group.addLayout(unten_group)
        ga_v.addLayout(button_group)

        # Canvas: zeigt Population (mehrere kleine Layouts)
        self.pop_canvas = PopulationCanvas(self, cols=10, rows=1)
        root.addWidget(self.pop_canvas)

        #=====================================================================================================
        #======================== Graphen der Fitness Funktion ===============================================
        
        self.graph_widget = pyqtgraph.PlotWidget()
        self.graph_widget.enableAutoRange(x = False, y = True)  # Automatische Anpassung der y-Achse
        self.graph_widget.setMouseEnabled(x=False, y=False)   # kein Drag/Pan, kein Wheel-Zoom
        self.graph_widget.setXRange(0, self.generations_spin.value(), padding=0.05)
        self.generations_spin.valueChanged.connect(self.on_generations_changed)

        self.fitness_Point_X: list[int] = []
        self.fitness_Point_Y: list[float] = []

        self.curve = self.graph_widget.plot(self.fitness_Point_X, self.fitness_Point_Y,pen=pyqtgraph.mkPen("#2D7FF9", width=1))

        self.graph_widget.setTitle("Fitness Verlauf", color="#000000", size="12pt", bold=True)
        self.graph_widget.setBackground("w")
        self.graph_widget.setLabel('left', 'Fitness',color="#000000", size="12pt", bold=True)
        self.graph_widget.setLabel('bottom', 'Generation',color="#000000", size="12pt", bold=True)
        root.addWidget(self.graph_widget)

        #=====================================================================================================

        # Status: globaler Status (Excel/GA Lauf etc.)
        self.status_label = QLabel("Bitte Excel Layout laden")
        root.addWidget(self.status_label)

        # Aktualisiert config.GRID_COLS/GRID_ROWS anhand config.FLOOR_* und config.GRID_SIZE
        update_grid_counts()

    @pyqtSlot(int)
    def on_generations_changed(self, value: int) -> None:
        max_gen = int(value)
        self.gencounter_label.setText(f"0 / {max_gen}")
        self.graph_widget.setXRange(0, max_gen, padding=0)

    # Öffnet Dateidialog, speichert Pfad und triggert Sheet-Liste aus Excel
    def load_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Excel Layout laden", "", "Excel files (*.xlsx *.xlsm *.xls)")
        if not path:
            return
        self.xlsx_path = path
        self._populate_sheets(path)

    # Liest Sheetnamen aus Excel, füllt das Dropdown, lädt optional erstes Sheet
    def _populate_sheets(self, path: str):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            sheets = list(wb.sheetnames)
        except Exception as e:
            print("ui_main: 261")
            QMessageBox.critical(self, "Excel Fehler", f"Konnte Excel nicht öffnen\n{e}")
            return

        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        for s in sheets:
            self.sheet_combo.addItem(s)
        self.sheet_combo.blockSignals(False)

        self.sheet_combo.setEnabled(True)
        self.reload_excel_btn.setEnabled(True)
        self.excel_file_label.setText(path)

        if sheets:
            self._apply_excel(path, sheets[0])

    # Lädt das aktuell ausgewählte Sheet erneut (z.B. nach Excel-Änderungen)
    def reload_excel(self):
        config.WATER_CELL = (None, None)
        config.GAS_CELL = (None, None)
        config.OTHER_CELL = (None, None)
        if self.xlsx_path and self.sheet_combo.currentText():
            self._apply_excel(self.xlsx_path, self.sheet_combo.currentText())

    # Event-Handler: wenn Sheet gewechselt wird, Layout neu laden
    def _on_sheet_changed(self, sheet: str):
        config.WATER_CELL = (None, None)
        config.GAS_CELL = (None, None)
        config.OTHER_CELL = (None, None)
        if self.xlsx_path and sheet:
            self._apply_excel(self.xlsx_path, sheet)

    # Lädt Excel-Daten ins config (Maschinen, Ports, Obstacles, Entry/Exit, Connections) und refresht UI
    def _apply_excel(self, path: str, sheet: str):
        try:
            apply_excel_layout_to_config(path, sheet_name=sheet)
            self.layout_loaded = True
        except Exception as e:
            self.layout_loaded = False
            print("ui_main: 295")
            QMessageBox.warning(self, "Excel Parse Fehler", str(e))
            return

        self.machine_info_label.setText(f"Maschinen {config.MACHINE_COUNT}")
        self.status_label.setText(
            f"Layout geladen {sheet} Floor {config.FLOOR_W:.2f} x {config.FLOOR_H:.2f} "
            f"Obstacles {len(config.OBSTACLES)} Flows {len(config.MATERIAL_CONNECTIONS)}"
        )

        self.pop_canvas.update()
        QApplication.processEvents()

    # Guard: ohne erfolgreich geladenes Layout wird GA nicht gestartet
    def _ensure_layout_loaded(self) -> bool:
        if self.layout_loaded:
            return True
        QMessageBox.information(self, "Layout fehlt", "Bitte zuerst ein Excel Layout laden")
        return False

    # Startet den GA synchron, aktualisiert UI via Callback, zeigt BestDialog am Ende
    def start_ga(self):
        if not self._ensure_layout_loaded():
            return

        #Übernimmt UI-Werte in config
        config.POPULATION_SIZE = int(self.population_size_spin.value())
        config.ELITE_KEEP = int(self.elite_keep_spin.value())
        config.MUTATION_PROB = float(self.mutation_prob_spin.value())
        config.MUTATION_POS_STD = float(self.mutation_pos_std_spin.value())
        config.MUTATION_ROT_PROB = float(self.mutation_rot_prob_spin.value())
        config.SWAP_PROB = float(self.mutation_swap_prob_spin.value())
        config.GENERATIONS = int(self.generations_spin.value())

        self.fitness_Point_X.clear()
        self.fitness_Point_Y.clear()    
        self.curve.setData(self.fitness_Point_X, self.fitness_Point_Y)

        gens = int(self.generations_spin.value())

        self.start_btn.setEnabled(False)
        self.status_label.setText("GA läuft")
        QApplication.processEvents()

        # Stop-Button an aktuellen Stop-Handler binden
        config.STOP_REQUESTED = False
        self.stop_btn.setEnabled(True)
        try:
            self.stop_btn.clicked.disconnect()
        except Exception:
            pass
        self.stop_btn.clicked.connect(self._stop)

        # Callback vom GA: UI-Labels + PopulationCanvas live aktualisieren
        def _cb(g, total, best_score_cb, best_ind_cb, population=None):
            self.gencounter_label.setText(f"{g} / {total}")
            if population:
                self.pop_canvas.set_population(population)
            self.status_label.setText(f"Gen {g} bester Score {best_score_cb:.2f}")
            self.fitness_Point_X.append(int(g))
            self.fitness_Point_Y.append(float(best_score_cb))
            self.curve.setData(self.fitness_Point_X, self.fitness_Point_Y)
            QApplication.processEvents()

        best_ind, best_score = run_ga(gens, progress_callback=_cb)
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)

        # Ergebnisdialog: zeigt beste Lösung, ansonsten Info-Dialog
        if best_ind:
            self._last_best_ind = best_ind
            self._last_best_score = best_score
            self.view_best_btn.setEnabled(True)
            BestDialog(best_ind, parent=self, title="Beste Lösung").exec()
        else:
            QMessageBox.information(self, "GA beendet", f"Kein Ergebnis Best score {best_score}")

    def view_best_result(self):
        """Öffnet das zuletzt gefundene beste Layout erneut (falls vorhanden)"""
        if not self._last_best_ind:
            QMessageBox.information(self, "Kein Ergebnis", "Es wurde noch kein bestes Ergebnis berechnet")
            return
        BestDialog(copy.deepcopy(self._last_best_ind), parent=self, title="Beste Lösung (gespeichert)").exec()

    # Stoppt den GA über ein globales Flag, das in run_ga periodisch geprüft wird
    def _stop(self):
        config.STOP_REQUESTED = True
        self.status_label.setText("Stop angefordert")
        QApplication.processEvents()