# -------------------------
# UI MAIN MODULE
# -------------------------
# Enthält das Hauptfenster (MainWindow)
#
# Änderungen:
# - Factory Editor entfernt (Hindernisse kommen aus Excel)
# - Bearbeitung der Maschinengrößen entfernt (Größen kommen aus Excel)
# - Layout wird über Excel geladen (Wände/Floor, Säulen -> Obstacles, Maschinen -> Sizes + Rotation)

from __future__ import annotations

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QLabel,
    QPushButton, QSpinBox, QDoubleSpinBox, QMessageBox, QApplication,
    QFileDialog, QComboBox
)
from PyQt6.QtCore import Qt

import config
from helpers import update_grid_counts
from ga_engine import GAEngine, run_ga
from ui_canvas import PopulationCanvas, BestDialog
from excel_import import apply_excel_layout_to_config


class MainWindow(QWidget):
    """Hauptfenster der Anwendung."""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Fabrikplaner GA")
        self.setMinimumSize(1024, 768)

        self.layout_loaded = False
        self.xlsx_path: str | None = None

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        # --- Steuerungsbereich (oben)
        controls_layout = QHBoxLayout()
        self.layout.addLayout(controls_layout)

        # --- Layout/Excel (links)
        layout_group = QGroupBox("Layout (Excel)")
        controls_layout.addWidget(layout_group)
        layout_v = QVBoxLayout()
        layout_group.setLayout(layout_v)

        self.excel_file_label = QLabel("Keine Excel-Datei geladen")
        self.sheet_combo = QComboBox()
        self.sheet_combo.setEnabled(False)
        self.sheet_combo.currentTextChanged.connect(self._on_sheet_changed)

        self.load_excel_btn = QPushButton("Excel laden…")
        self.load_excel_btn.clicked.connect(self.load_excel)

        self.reload_excel_btn = QPushButton("Layout neu laden")
        self.reload_excel_btn.clicked.connect(self.reload_excel)
        self.reload_excel_btn.setEnabled(False)

        self.machine_info_label = QLabel(f"Maschinen: {config.MACHINE_COUNT}")

        layout_v.addWidget(self.load_excel_btn)
        layout_v.addWidget(self.reload_excel_btn)
        layout_v.addWidget(QLabel("Sheet:"))
        layout_v.addWidget(self.sheet_combo)
        layout_v.addWidget(self.excel_file_label)
        layout_v.addWidget(self.machine_info_label)

        # --- Entry / Exit Controls
        entry_group = QGroupBox("Entry / Exit (Zellen)")
        controls_layout.addWidget(entry_group)
        entry_layout = QVBoxLayout()
        entry_group.setLayout(entry_layout)

        from PyQt6.QtWidgets import QGridLayout
        entry_grid = QGridLayout()
        entry_layout.addLayout(entry_grid)

        self.entry_col = QSpinBox()
        self.entry_row = QSpinBox()
        self.exit_col = QSpinBox()
        self.exit_row = QSpinBox()
        self._refresh_entry_exit_ranges()

        entry_grid.addWidget(QLabel("Entry Col"), 0, 0)
        entry_grid.addWidget(self.entry_col, 0, 1)
        entry_grid.addWidget(QLabel("Entry Row"), 0, 2)
        entry_grid.addWidget(self.entry_row, 0, 3)
        entry_grid.addWidget(QLabel("Exit Col"), 1, 0)
        entry_grid.addWidget(self.exit_col, 1, 1)
        entry_grid.addWidget(QLabel("Exit Row"), 1, 2)
        entry_grid.addWidget(self.exit_row, 1, 3)

        self.entry_col.valueChanged.connect(self.on_entry_exit_changed)
        self.entry_row.valueChanged.connect(self.on_entry_exit_changed)
        self.exit_col.valueChanged.connect(self.on_entry_exit_changed)
        self.exit_row.valueChanged.connect(self.on_entry_exit_changed)

        # --- Rechte Seite: GA-Parameter
        ga_group = QGroupBox("Genetischer Algorithmus")
        controls_layout.addWidget(ga_group)
        ga_layout = QVBoxLayout()
        ga_group.setLayout(ga_layout)

        self.population_size_spin = QSpinBox()
        self.population_size_spin.setRange(1, 1000)
        self.population_size_spin.setValue(config.POPULATION_SIZE)
        ga_layout.addWidget(QLabel("Population Größe:"))
        ga_layout.addWidget(self.population_size_spin)

        self.elite_keep_spin = QSpinBox()
        self.elite_keep_spin.setRange(1, 1000)
        self.elite_keep_spin.setValue(config.ELITE_KEEP)
        ga_layout.addWidget(QLabel("Eliten behalten:"))
        ga_layout.addWidget(self.elite_keep_spin)

        self.mutation_prob_spin = QDoubleSpinBox()
        self.mutation_prob_spin.setRange(0.0, 1.0)
        self.mutation_prob_spin.setValue(config.BASE_MUTATION_PROB)
        self.mutation_prob_spin.setSingleStep(0.01)
        ga_layout.addWidget(QLabel("Mutationswahrscheinlichkeit:"))
        ga_layout.addWidget(self.mutation_prob_spin)

        self.mutation_pos_std_spin = QDoubleSpinBox()
        self.mutation_pos_std_spin.setRange(0.1, 10.0)
        self.mutation_pos_std_spin.setValue(config.BASE_MUTATION_POS_STD)
        self.mutation_pos_std_spin.setSingleStep(0.1)
        ga_layout.addWidget(QLabel("Positions-Mutations-StdAbw.:"))
        ga_layout.addWidget(self.mutation_pos_std_spin)

        # Rotation Mutation bleibt aus UI sichtbar, wirkt aber nicht (Rotation wird fix gehalten)
        self.mutation_angle_std_spin = QDoubleSpinBox()
        self.mutation_angle_std_spin.setRange(0.1, 10.0)
        self.mutation_angle_std_spin.setValue(config.BASE_MUTATION_ANGLE_STD)
        self.mutation_angle_std_spin.setSingleStep(0.1)
        ga_layout.addWidget(QLabel("Rotations-Mutations-StdAbw.: (derzeit deaktiv)"))
        ga_layout.addWidget(self.mutation_angle_std_spin)

        self.generations_spin = QSpinBox()
        self.generations_spin.setRange(1, 10000)
        self.generations_spin.setValue(200)
        ga_layout.addWidget(QLabel("Generationen (Total):"))
        ga_layout.addWidget(self.generations_spin)

        self.start_btn = QPushButton("GA starten (vollständig)")
        self.start_btn.clicked.connect(self.start_ga)
        ga_layout.addWidget(self.start_btn)

        self.stop_btn = QPushButton("Abbrechen")
        self.stop_btn.setEnabled(False)
        ga_layout.addWidget(self.stop_btn)

        # --- Generations-per-click und Next-Button
        adv_layout = QHBoxLayout()
        ga_layout.addLayout(adv_layout)
        self.advance_spin = QSpinBox()
        self.advance_spin.setRange(1, 1000)
        self.advance_spin.setValue(1)
        adv_layout.addWidget(QLabel("Generationen / Klick:"))
        adv_layout.addWidget(self.advance_spin)

        self.next_gen_btn = QPushButton("Nächste Generationen")
        self.next_gen_btn.clicked.connect(self.on_next_generation)
        adv_layout.addWidget(self.next_gen_btn)

        self.show_best_spin = QSpinBox()
        self.show_best_spin.setRange(0, 1000)
        self.show_best_spin.setValue(10)
        adv_layout.addWidget(QLabel("Bestanzeige alle N Gen.:"))
        adv_layout.addWidget(self.show_best_spin)

        self.gencounter_label = QLabel("0/0")
        ga_layout.addWidget(self.gencounter_label)

        # --- Zentraler Bereich: Population-Canvas
        self.pop_canvas = PopulationCanvas(self, cols=10, rows=5)
        self.layout.addWidget(self.pop_canvas)

        # --- Statusleiste (unten)
        self.status_label = QLabel("Bitte Excel-Layout laden.")
        self.layout.addWidget(self.status_label)

        self.ga_engine = None

        # ensure grid counts are up-to-date at startup
        update_grid_counts()
        self._refresh_entry_exit_ranges()

    def closeEvent(self, event):
        reply = QMessageBox.question(
            self,
            "Bestätigung",
            "Möchten Sie das Programm wirklich beenden?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            event.accept()
        else:
            event.ignore()

    # -------------------------
    # Excel-Import
    # -------------------------
    def load_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Excel Layout laden", "", "Excel files (*.xlsx *.xlsm *.xls)")
        if not path:
            return
        self.xlsx_path = path
        self._populate_sheets(path)

    def _populate_sheets(self, path: str):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            sheets = list(wb.sheetnames)
        except Exception as e:
            QMessageBox.critical(self, "Excel Fehler", f"Konnte Excel nicht öffnen:\n{e}")
            return

        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        for s in sheets:
            self.sheet_combo.addItem(s)
        self.sheet_combo.blockSignals(False)
        self.sheet_combo.setEnabled(True)
        self.reload_excel_btn.setEnabled(True)
        self.excel_file_label.setText(path)

        if sheets:
            self._apply_excel(path, sheets[0])

    def reload_excel(self):
        if not self.xlsx_path or not self.sheet_combo.currentText():
            return
        self._apply_excel(self.xlsx_path, self.sheet_combo.currentText())

    def _on_sheet_changed(self, sheet: str):
        if not self.xlsx_path or not sheet:
            return
        self._apply_excel(self.xlsx_path, sheet)

    def _apply_excel(self, path: str, sheet: str):
        try:
            apply_excel_layout_to_config(path, sheet_name=sheet)
            self.layout_loaded = True
        except Exception as e:
            self.layout_loaded = False
            QMessageBox.warning(self, "Excel Parse-Fehler", str(e))
            return

        self.machine_info_label.setText(f"Maschinen: {config.MACHINE_COUNT}")
        self.status_label.setText(
            f"Layout geladen: {sheet} — Floor: {config.FLOOR_W:.2f}m x {config.FLOOR_H:.2f}m — Obstacles: {len(config.OBSTACLES)}"
        )

        self._refresh_entry_exit_ranges()
        self.pop_canvas.update()
        QApplication.processEvents()

    def _refresh_entry_exit_ranges(self):
        self.entry_col.setRange(0, max(0, config.GRID_COLS - 1))
        self.exit_col.setRange(0, max(0, config.GRID_COLS - 1))
        self.entry_row.setRange(0, max(0, config.GRID_ROWS - 1))
        self.exit_row.setRange(0, max(0, config.GRID_ROWS - 1))
        self.entry_col.setValue(int(config.ENTRY_CELL[0]))
        self.entry_row.setValue(int(config.ENTRY_CELL[1]))
        self.exit_col.setValue(int(config.EXIT_CELL[0]))
        self.exit_row.setValue(int(config.EXIT_CELL[1]))

    def on_entry_exit_changed(self):
        config.ENTRY_CELL = (int(self.entry_col.value()), int(self.entry_row.value()))
        config.EXIT_CELL = (int(self.exit_col.value()), int(self.exit_row.value()))
        self.status_label.setText(f"Entry/Exit gesetzt: {config.ENTRY_CELL} -> {config.EXIT_CELL}")
        self.pop_canvas.update()

    # -------------------------
    # GA Steuerung
    # -------------------------
    def _ensure_layout_loaded(self) -> bool:
        if self.layout_loaded:
            return True
        QMessageBox.information(self, "Layout fehlt", "Bitte zuerst ein Excel-Layout laden.")
        return False

    def on_next_generation(self):
        if not self._ensure_layout_loaded():
            return

        if self.ga_engine is None:
            total = int(self.generations_spin.value())
            self.ga_engine = GAEngine(total)
            self.pop_canvas.set_population(self.ga_engine.population)
            self.gencounter_label.setText(f"{self.ga_engine.generation}/{self.ga_engine.total_generations}")

        advance = int(self.advance_spin.value())
        show_every = int(self.show_best_spin.value())

        for _ in range(advance):
            if self.ga_engine.generation >= self.ga_engine.total_generations:
                break
            best_score, _best_ind = self.ga_engine.step()
            self.pop_canvas.set_population(self.ga_engine.population)
            self.pop_canvas.update()
            self.gencounter_label.setText(f"{self.ga_engine.generation}/{self.ga_engine.total_generations}")
            self.status_label.setText(f"Gen {self.ga_engine.generation}: bester Score {best_score:.2f}")
            QApplication.processEvents()

            if show_every > 0 and (self.ga_engine.generation % show_every == 0):
                if self.ga_engine.best_ind:
                    dialog = BestDialog(
                        self.ga_engine.best_ind,
                        parent=self,
                        title=f"Beste Lösung nach Generation {self.ga_engine.generation}",
                    )
                    dialog.exec()

        if self.ga_engine.generation >= self.ga_engine.total_generations:
            if self.ga_engine.best_ind:
                dialog = BestDialog(
                    self.ga_engine.best_ind,
                    parent=self,
                    title=f"Beste Lösung (Ende Gen {self.ga_engine.generation})",
                )
                dialog.exec()
            self.next_gen_btn.setEnabled(False)
            self.start_btn.setEnabled(True)
            self.stop_btn.setEnabled(False)

    def start_ga(self):
        if not self._ensure_layout_loaded():
            return

        config.POPULATION_SIZE = int(self.population_size_spin.value())
        config.ELITE_KEEP = int(self.elite_keep_spin.value())
        config.MUTATION_PROB = float(self.mutation_prob_spin.value())
        config.MUTATION_POS_STD = float(self.mutation_pos_std_spin.value())
        config.MUTATION_ANGLE_STD = float(self.mutation_angle_std_spin.value())
        gens = int(self.generations_spin.value())

        # UI sperren
        self.start_btn.setEnabled(False)
        self.next_gen_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.status_label.setText("GA läuft...")
        QApplication.processEvents()

        config.STOP_REQUESTED = False
        try:
            self.stop_btn.clicked.disconnect()
        except Exception:
            pass
        self.stop_btn.clicked.connect(lambda: self._set_stop_flag())

        def _cb(generation, total_generations, best_score_cb, best_ind_cb, population=None):
            try:
                self._on_worker_progress(generation, total_generations, best_score_cb, best_ind_cb, population)
            except Exception:
                pass

        try:
            best_ind, best_score = run_ga(gens, progress_callback=_cb)
        except Exception:
            best_ind, best_score = None, float("inf")

        self._on_worker_finished(best_ind, best_score)

    def _set_stop_flag(self):
        config.STOP_REQUESTED = True
        self.status_label.setText("Stop angefordert...")
        QApplication.processEvents()

    def _on_worker_progress(self, generation, total_generations, best_score, best_ind, population):
        try:
            if population:
                self.pop_canvas.set_population(population)
            elif best_ind is not None:
                self.pop_canvas.set_population([best_ind])
            self.gencounter_label.setText(f"{generation}/{total_generations}")
            self.status_label.setText(f"Gen {generation}/{total_generations} — bester Score: {best_score:.2f}")
            QApplication.processEvents()
        except Exception:
            pass

    def _on_worker_finished(self, best_ind, best_score):
        try:
            if best_ind:
                dialog = BestDialog(best_ind, parent=self, title="Beste Lösung (Ende)")
                dialog.exec()
            else:
                QMessageBox.information(self, "GA beendet", f"GA abgebrochen oder kein Ergebnis. Best score: {best_score}")
        except Exception:
            pass

        self.start_btn.setEnabled(True)
        self.next_gen_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
