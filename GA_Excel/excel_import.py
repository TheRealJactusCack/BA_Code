# -------------------------
# EXCEL IMPORT MODULE
# -------------------------
# Liest Layout-Daten (Wände, Säulen/Hindernisse, Maschinen) aus einer Excel-Datei
# und schreibt sie in `config.*` um den GA ohne manuelle Eingaben zu starten.
#
# Benötigt: openpyxl
#
# Erwartetes Tabellenformat (ähnlich zu c03e29f0-...py):
#   - Oberhalb der Header-Zeile optional Meta-Infos in Spalte A/B: key | value
#   - Header-Zeile enthält mindestens: id | type
#   - Danach Datenzeilen:
#       type == "wall":   x1, y1, x2, y2
#       type == "column": x, y, w, d, rot
#       type == "machine": w, d, input   (input optional; wird auf 0/90/180/270° gerundet)
#
# Koordinatensystem:
#   Excel wird als kartesisch (y nach oben) interpretiert.
#   Intern (Canvas/GA) wird y nach unten benutzt.
#   Daher wird y standardmäßig geflippt: y_internal = (max_y - y_excel).
#
# Optionales Meta (key in A, value in B):
#   - grid_size:   Rastergröße in Metern (float)
#   - floor_w / floor_h: explizite Floor-Größe (float). Wenn nicht gesetzt:
#                         Bounding-Box aus Wänden wird genutzt.
#   - flip_y:      true/false (default true)
#   - entry_col, entry_row, exit_col, exit_row: Entry/Exit in Zellen
#   - entry_x, entry_y, exit_x, exit_y: Entry/Exit in Metern (Excel-Koordinaten)

from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import math

from openpyxl import load_workbook

import config
from helpers import update_grid_counts


@dataclass(frozen=True)
class Wall:
    x1: float
    y1: float
    x2: float
    y2: float


@dataclass(frozen=True)
class Column:
    x: float
    y: float
    w: float
    d: float
    rot: float = 0.0


@dataclass(frozen=True)
class MachineRow:
    id: str
    w: float
    d: float
    input: float = 0.0


def _f(v: Any) -> Optional[float]:
    try:
        return None if v is None else float(v)
    except Exception:
        return None


def _truthy(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in {"1", "true", "yes", "y", "ja", "on"}


def _nearest_rotation_deg(deg: float) -> int:
    """Rundet auf 0/90/180/270."""
    opts = [0, 90, 180, 270]
    d = float(deg) % 360.0
    return min(opts, key=lambda a: abs(((d - a + 180) % 360) - 180))


def read_layout_from_sheet(xlsx_path: str, sheet_name: str) -> Tuple[List[Wall], List[Column], List[MachineRow], Dict[str, Any]]:
    """
    Parst Excel wie im externen Viewer.

    Returns:
        walls, columns, machines, meta
    """
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]

    header_row = None
    for r in range(1, 31):
        a = str(ws.cell(r, 1).value or "").strip().lower()
        b = str(ws.cell(r, 2).value or "").strip().lower()
        if a == "id" and b == "type":
            header_row = r
            break
    if header_row is None:
        raise ValueError("Header (id,type) nicht gefunden")

    header_map: Dict[str, int] = {}
    for c in range(1, 80):
        val = ws.cell(header_row, c).value
        if val is None:
            break
        header_map[str(val).strip().lower()] = c

    def idx(name: str) -> Optional[int]:
        return header_map.get(name)

    meta: Dict[str, Any] = {}
    for r in range(1, header_row):
        k = ws.cell(r, 1).value
        if k is None:
            continue
        meta[str(k).strip().lower()] = ws.cell(r, 2).value

    walls: List[Wall] = []
    cols: List[Column] = []
    machines: List[MachineRow] = []

    blank = 0
    for r in range(header_row + 1, ws.max_row + 1):
        raw_id = ws.cell(r, idx("id") or 0).value
        if raw_id is None or str(raw_id).strip() == "":
            blank += 1
            if blank >= 5:
                break
            continue
        blank = 0

        t = str(ws.cell(r, idx("type") or 0).value or "").strip().lower()

        if t == "wall":
            x1 = _f(ws.cell(r, idx("x1") or 0).value)
            y1 = _f(ws.cell(r, idx("y1") or 0).value)
            x2 = _f(ws.cell(r, idx("x2") or 0).value)
            y2 = _f(ws.cell(r, idx("y2") or 0).value)
            if None not in (x1, y1, x2, y2):
                walls.append(Wall(x1, y1, x2, y2))

        elif t == "column":
            x = _f(ws.cell(r, idx("x") or 0).value)
            y = _f(ws.cell(r, idx("y") or 0).value)
            if x is None or y is None:
                continue
            w = _f(ws.cell(r, idx("w") or 0).value) or 0.5
            d = _f(ws.cell(r, idx("d") or 0).value) or 0.5
            rot = _f(ws.cell(r, idx("rot") or 0).value) or 0.0
            cols.append(Column(x=x, y=y, w=float(w), d=float(d), rot=float(rot)))

        elif t == "machine":
            label = str(raw_id)
            w = _f(ws.cell(r, idx("w") or 0).value) or 1.0
            d = _f(ws.cell(r, idx("d") or 0).value) or 1.0
            inputv = _f(ws.cell(r, idx("input") or 0).value) or 0.0
            machines.append(MachineRow(id=label, w=float(w), d=float(d), input=float(inputv)))

    return walls, cols, machines, meta


def _compute_bounds_from_walls(walls: List[Wall]) -> Optional[Tuple[float, float, float, float]]:
    if not walls:
        return None
    xs = [v for w in walls for v in (w.x1, w.x2)]
    ys = [v for w in walls for v in (w.y1, w.y2)]
    if not xs or not ys:
        return None
    return min(xs), max(xs), min(ys), max(ys)


def _point_in_rotated_rect(px: float, py: float, cx: float, cy: float, w: float, h: float, rot_deg: float) -> bool:
    """Testet Punkt gegen rot. Rechteck (rot um center)."""
    # rotate point into rect-local coords
    ang = math.radians(rot_deg)
    cos_a = math.cos(-ang)
    sin_a = math.sin(-ang)
    dx = px - cx
    dy = py - cy
    lx = dx * cos_a - dy * sin_a
    ly = dx * sin_a + dy * cos_a
    return (abs(lx) <= w / 2.0 + 1e-9) and (abs(ly) <= h / 2.0 + 1e-9)


def _cells_for_rotated_rect(cx: float, cy: float, w: float, h: float, rot_deg: float) -> set[Tuple[int, int]]:
    """Approximiert belegte Rasterzellen über Zellzentren."""
    # corners for bbox
    ang = math.radians(rot_deg)
    dx = w / 2.0
    dy = h / 2.0
    corners = [(-dx, -dy), (dx, -dy), (dx, dy), (-dx, dy)]
    cos_a = math.cos(ang)
    sin_a = math.sin(ang)
    pts = []
    for x, y in corners:
        rx = x * cos_a - y * sin_a + cx
        ry = x * sin_a + y * cos_a + cy
        pts.append((rx, ry))
    minx = min(p[0] for p in pts)
    maxx = max(p[0] for p in pts)
    miny = min(p[1] for p in pts)
    maxy = max(p[1] for p in pts)

    gs = float(config.GRID_SIZE)
    c0 = max(0, int(math.floor(minx / gs)))
    c1 = min(config.GRID_COLS - 1, int(math.floor(maxx / gs)))
    r0 = max(0, int(math.floor(miny / gs)))
    r1 = min(config.GRID_ROWS - 1, int(math.floor(maxy / gs)))

    out: set[Tuple[int, int]] = set()
    for col in range(c0, c1 + 1):
        for row in range(r0, r1 + 1):
            px = (col + 0.5) * gs
            py = (row + 0.5) * gs
            if _point_in_rotated_rect(px, py, cx, cy, w, h, rot_deg):
                out.add((col, row))
    return out


def apply_excel_layout_to_config(xlsx_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    Lädt Excel-Layout und schreibt nach `config`.

    Returns:
        meta dict (bereits lowercased keys)
    """
    wb = load_workbook(xlsx_path, data_only=True)
    if not wb.sheetnames:
        raise ValueError("Workbook hat keine Sheets")
    sheet = sheet_name or wb.sheetnames[0]

    walls, cols, machines, meta = read_layout_from_sheet(xlsx_path, sheet)

    # grid size (optional)
    if "grid_size" in meta and _f(meta["grid_size"]) is not None:
        config.GRID_SIZE = float(_f(meta["grid_size"]))

    # bounds / floor size
    bounds = _compute_bounds_from_walls(walls)
    if bounds is not None:
        minx, maxx, miny, maxy = bounds
        floor_w = float(maxx - minx)
        floor_h = float(maxy - miny)
    else:
        minx = 0.0
        miny = 0.0
        floor_w = float(_f(meta.get("floor_w")) or config.FLOOR_W)
        floor_h = float(_f(meta.get("floor_h")) or config.FLOOR_H)
        maxy = miny + floor_h

    # override via meta if given
    mw = _f(meta.get("floor_w"))
    mh = _f(meta.get("floor_h"))
    if mw is not None:
        floor_w = float(mw)
    if mh is not None:
        floor_h = float(mh)

    config.FLOOR_W = max(1e-6, float(floor_w))
    config.FLOOR_H = max(1e-6, float(floor_h))

    update_grid_counts()

    # coordinate transform excel -> internal
    flip_y = True if "flip_y" not in meta else _truthy(meta.get("flip_y"))
    maxy_for_flip = maxy if bounds is not None else (miny + floor_h)

    def to_internal_xy(x: float, y: float) -> Tuple[float, float]:
        x0 = float(x) - float(minx)
        if flip_y:
            y0 = float(maxy_for_flip) - float(y)
        else:
            y0 = float(y) - float(miny)
        return x0, y0

    # machines: sizes + fixed rotations
    if machines:
        config.MACHINE_COUNT = int(len(machines))
        config.MACHINE_SIZES = [(max(0.1, m.w), max(0.1, m.d)) for m in machines]
        config.MACHINE_ROTATIONS = [_nearest_rotation_deg(m.input) for m in machines]
        config.MACHINE_LABELS = [m.id for m in machines]
    else:
        # keep existing defaults but ensure list size
        config.MACHINE_SIZES = list(config.MACHINE_SIZES[: config.MACHINE_COUNT])

    # obstacles from columns
    obstacles: set[Tuple[int, int]] = set()
    for c in cols:
        cx, cy = to_internal_xy(c.x, c.y)
        w = max(0.05, float(c.w))
        h = max(0.05, float(c.d))
        rot = float(c.rot)
        obstacles |= _cells_for_rotated_rect(cx, cy, w, h, rot)

    # clamp to grid
    obstacles = {cell for cell in obstacles if 0 <= cell[0] < config.GRID_COLS and 0 <= cell[1] < config.GRID_ROWS}
    config.OBSTACLES = obstacles

    # entry/exit
    def set_entry_exit_from_meta() -> None:
        # cells explicit
        ec = meta.get("entry_col")
        er = meta.get("entry_row")
        xc = meta.get("exit_col")
        xr = meta.get("exit_row")
        if _f(ec) is not None and _f(er) is not None:
            config.ENTRY_CELL = (int(_f(ec)), int(_f(er)))
        if _f(xc) is not None and _f(xr) is not None:
            config.EXIT_CELL = (int(_f(xc)), int(_f(xr)))

        # meters
        exx = _f(meta.get("entry_x"))
        exy = _f(meta.get("entry_y"))
        if exx is not None and exy is not None:
            ix, iy = to_internal_xy(exx, exy)
            config.ENTRY_CELL = (int(ix // config.GRID_SIZE), int(iy // config.GRID_SIZE))

        outx = _f(meta.get("exit_x"))
        outy = _f(meta.get("exit_y"))
        if outx is not None and outy is not None:
            ix, iy = to_internal_xy(outx, outy)
            config.EXIT_CELL = (int(ix // config.GRID_SIZE), int(iy // config.GRID_SIZE))

    set_entry_exit_from_meta()
    update_grid_counts()  # clamps entry/exit

    return meta
