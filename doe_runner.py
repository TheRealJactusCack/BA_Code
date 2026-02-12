# FILE: doe_runner.py
"""
DOE für den GA

Eigenständig und unabhängig startbar:
  python doe_runner.py

Was passiert:
  - Excel + Sheet auswählen (Layout laden)
  - DOE-Parameterbereiche definieren (Config-Keys)
  - GA mehrfach laufen lassen (Design-Punkte)
  - Ergebnisse tabellarisch anzeigen + BestLayout pro Run ansehen

Hinweis:
  - DOE Stop nutzt config.STOP_REQUESTED (wie im Main-UI)
"""

from __future__ import annotations

import atexit
import itertools
import math
import sys
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import load_workbook
from PyQt6.QtCore import QObject, Qt, QThread, pyqtSignal, QCoreApplication
from PyQt6.QtWidgets import (QApplication, QComboBox, QDoubleSpinBox, QFileDialog,
    QGroupBox, QHBoxLayout, QLabel, QMessageBox, QPushButton, QSpinBox, QTableWidget,
    QTableWidgetItem, QVBoxLayout, QWidget, QHeaderView
)
import config
from excel_import import apply_excel_layout_to_config
from ga_engine import run_ga
from ui_canvas import BestDialog

@dataclass(frozen=True)
class DOEParam:
    """Beschreibt einen DOE-Parameter als Bereich (Start/Stop) mit einer Anzahl an tests/ Parameteränderungen (Steps)."""

    key: str
    start: float
    stop: float
    steps: int
    kind: str  # "int" or "float"


# Erzeugt n gleichmäßig verteilte Werte zwischen a und b (inkl. a und b) (alle werte für die Steps)
def _linspace(a: float, b: float, n: int) -> List[float]:
    if n <= 1:
        return [float(a)]
    if math.isclose(a, b):
        return [float(a)] * n
    step = (b - a) / (n - 1)
    return [a + i * step for i in range(n)]


# Castet einen float-Wert abhängig vom Parameter-Typ (int/float) in den Zieltyp
def _cast_value(v: float, kind: str) -> Any:
    if kind == "int":
        return int(round(v))
    return float(v)


# Baut aus den Parameterbereichen alle DOE-Designpunkte (Full-Factorial)
def _design_points(params: Sequence[DOEParam]) -> List[Dict[str, Any]]:
    grids: List[List[Any]] = []
    keys: List[str] = []
    for p in params:
        keys.append(p.key)
        values = [_cast_value(v, p.kind) for v in _linspace(p.start, p.stop, p.steps)]
        grids.append(values)

    points: List[Dict[str, Any]] = []
    for combo in itertools.product(*grids):
        points.append({k: combo[i] for i, k in enumerate(keys)})
    return points


# ------------------------------ Worker Thread ----------------------------------


class DOEWorker(QObject):
    """Worker, der DOE-Runs im Hintergrund ausführt und Ergebnisse per Signals zurückmeldet"""

    progressed = pyqtSignal(int, int, str)  # i, total, status
    result_ready = pyqtSignal(int, dict, float, object)  # run_idx, point, score, best_ind
    finished = pyqtSignal()

    def __init__(self, generations: int, points: List[Dict[str, Any]]) -> None:
        """Initialisiert den Worker mit GA-Generationen und den DOE-Designpunkten."""
        super().__init__()
        self.generations = int(generations)
        self.points = points
        self._stop_requested = False

    # Markiert den Worker als zu stoppen und setzt das globale Stop-Flag des GA
    def request_stop(self) -> None:
        self._stop_requested = True
        config.STOP_REQUESTED = True

    # Führt alle Designpunkte nacheinander aus: config setzen -> GA laufen lassen -> Ergebnis emitten
    def run(self) -> None:
        total = len(self.points)
        for i, point in enumerate(self.points, start=1):
            if self._stop_requested:
                break

            # Config setzen
            for k, v in point.items():
                if not hasattr(config, k):
                    self.progressed.emit(i, total, f"ERROR: config hat kein Attribut '{k}'")
                    self._stop_requested = True
                    break
                setattr(config, k, v)

            if self._stop_requested:
                break

            config.STOP_REQUESTED = False
            self.progressed.emit(i, total, f"Run {i}/{total} ...")

            best_ind, best_score = run_ga(self.generations, progress_callback=None)
            if best_ind is None:
                best_score = float("inf")

            self.result_ready.emit(i, point, float(best_score), best_ind)

        self.finished.emit()


# ---------------------------------- UI -----------------------------------------


class DOERunnerWindow(QWidget):
    """Eigenständiges DOE-Fenster: Excel wählen, DOE konfigurieren, DOE ausführen, Ergebnisse ansehen."""

    def __init__(self) -> None:
        """Baut die komplette UI und initialisiert State/Worker-Verwaltung."""
        super().__init__()
        self.setWindowTitle("DOE Runner (GA)")
        self.setMinimumSize(1200, 720)

        self.xlsx_path: Optional[str] = None
        self.layout_loaded = False

        self._worker_thread: Optional[QThread] = None
        self._worker: Optional[DOEWorker] = None

        # Speichert pro DOE-Run das zugehörige Best-Layout (ind).
        self._best_by_run: Dict[int, Any] = {}

        root = QVBoxLayout(self)

        # --- Excel group ---
        excel_group = QGroupBox("Excel / Layout")
        root.addWidget(excel_group)
        excel_layout = QHBoxLayout(excel_group)

        self.excel_label = QLabel("Keine Excel Datei geladen")
        self.sheet_combo = QComboBox()
        self.sheet_combo.setEnabled(False)
        self.sheet_combo.currentTextChanged.connect(self._on_sheet_changed)

        self.btn_load_excel = QPushButton("Excel laden")
        self.btn_load_excel.clicked.connect(self.load_excel)

        self.btn_reload_excel = QPushButton("Layout neu laden")
        self.btn_reload_excel.setEnabled(False)
        self.btn_reload_excel.clicked.connect(self.reload_excel)

        excel_layout.addWidget(self.btn_load_excel)
        excel_layout.addWidget(self.btn_reload_excel)
        excel_layout.addWidget(QLabel("Sheet"))
        excel_layout.addWidget(self.sheet_combo, 1)
        excel_layout.addWidget(self.excel_label, 3)

        # --- DOE parameter group ---
        doe_group = QGroupBox("DOE Parameter")
        root.addWidget(doe_group)
        doe_v = QVBoxLayout(doe_group)

        self.params_table = QTableWidget(0, 5)
        self.params_table.setHorizontalHeaderLabels(["Key", "Start", "Stop", "Steps", "Type"])
        header = self.params_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        btn_row = QHBoxLayout()
        self.btn_add_param = QPushButton("+ Parameter")
        self.btn_add_param.clicked.connect(self.add_param_row)
        self.btn_del_param = QPushButton("- Parameter")
        self.btn_del_param.clicked.connect(self.del_param_row)

        btn_row.addWidget(self.btn_add_param)
        btn_row.addWidget(self.btn_del_param)
        btn_row.addStretch(1)

        doe_v.addLayout(btn_row)
        doe_v.addWidget(self.params_table)

        #Defaults
        self._add_param_row_default("POPULATION_SIZE", 10, 200, 4, "int")
        self._add_param_row_default("GENERATIONS", 50, 1000, 4, "int")

        #GA settings / Durchlauf 
        run_group = QGroupBox("Durchlauf")
        root.addWidget(run_group)
        run_h = QHBoxLayout(run_group)
        """
        self.gens_spin = QSpinBox()
        self.gens_spin.setRange(1, 20000)
        self.gens_spin.setValue(300)"""

        self.btn_run = QPushButton("DOE Starten")
        self.btn_run.clicked.connect(self.run_doe)

        self.btn_stop = QPushButton("Stop")
        self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self.stop_doe)

        self.status = QLabel("Bereit")
       
        run_h.addWidget(QLabel("Generationen"))
        """
        run_h.addWidget(self.gens_spin)"""
        run_h.addWidget(self.btn_run)
        run_h.addWidget(self.btn_stop)
        run_h.addWidget(self.status, 1)

        # --- Results ---
        res_group = QGroupBox("Ergebnisse")
        root.addWidget(res_group, 1)
        res_v = QVBoxLayout(res_group)

        self.results_table = QTableWidget(0, 2)
        self.results_table.verticalHeader().setVisible(False)
        self.results_table.setHorizontalHeaderLabels(["Durchlauf", "Bester score"])
        self.results_table.horizontalHeader().setStretchLastSection(True)

        self.btn_view = QPushButton("View Selected Best Layout")
        self.btn_view.clicked.connect(self.view_selected_best)

        res_v.addWidget(self.results_table, 1)
        res_v.addWidget(self.btn_view)

    # Öffnet Dateidialog, liest Sheets aus Excel und lädt initial das erste Sheet
    def load_excel(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Excel auswählen", "", "Excel (*.xlsx *.xlsm)")
        if not path:
            return
        self.xlsx_path = path
        self.excel_label.setText(path)

        try:
            wb = load_workbook(path, data_only=True)
            sheets = wb.sheetnames
        except Exception as e:
            QMessageBox.critical(self, "Excel Fehler", f"Konnte Excel nicht öffnen:\n{e}")
            return

        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        for s in sheets:
            self.sheet_combo.addItem(s)
        self.sheet_combo.blockSignals(False)
        self.sheet_combo.setEnabled(True)
        self.btn_reload_excel.setEnabled(True)

        if sheets:
            self._apply_excel(path, sheets[0])

    # Lädt das aktuell ausgewählte Sheet erneut (z.B. nach Excel-Änderungen)
    def reload_excel(self) -> None:
        if self.xlsx_path and self.sheet_combo.currentText():
            self._apply_excel(self.xlsx_path, self.sheet_combo.currentText())

    # Event-Handler: Sheet-Wechsel triggert Layout-Neuladen
    def _on_sheet_changed(self, sheet: str) -> None:
        if self.xlsx_path and sheet:
            self._apply_excel(self.xlsx_path, sheet)

    # Übernimmt Excel/SHEET in config (MACHINE_*, MATERIAL_CONNECTIONS, ENTRY/EXIT etc.)
    def _apply_excel(self, path: str, sheet: str) -> None:
        try:
            apply_excel_layout_to_config(path, sheet_name=sheet)
            self.layout_loaded = True
        except Exception as e:
            self.layout_loaded = False
            QMessageBox.warning(self, "Excel Parse Fehler", str(e))
            print("doe_runner: 291")
            return

        self.status.setText(
            f"Layout geladen: {sheet} | Maschinen={config.MACHINE_COUNT} | Flows={len(getattr(config, 'MATERIAL_CONNECTIONS', []))}"
        )

    # Prüft, ob ein Layout geladen ist (sonst Hinweisdialog)
    def _ensure_layout_loaded(self) -> bool:
        if self.layout_loaded:
            return True
        QMessageBox.information(self, "Layout fehlt", "Bitte zuerst ein Excel Layout laden.")
        return False

    # Fügt eine Param-Zeile mit Widgets ein (Key, Start, Stop, Steps, Type).
    def _add_param_row_default(self, key: str, start: float, stop: float, steps: int, kind: str) -> None:
        row = self.params_table.rowCount()
        self.params_table.insertRow(row)

        key_combo = QComboBox()
        key_combo.addItems(
            [
                "POPULATION_SIZE",
                "GENERATIONS",
                "ELITE_KEEP",
                "MUTATION_PROB",
                "MUTATION_POS_STD",
                "MUTATION_ROT_PROB",
                "GRID_SIZE",
                "MACHINE_CLEARANCE_M",
            ]
        )
        if key in [key_combo.itemText(i) for i in range(key_combo.count())]:
            key_combo.setCurrentText(key)
        else:
            key_combo.addItem(key)
            key_combo.setCurrentText(key)
        self.params_table.setCellWidget(row, 0, key_combo)

        start_box = QDoubleSpinBox()
        start_box.setRange(-1e9, 1e9)
        start_box.setDecimals(2)
        start_box.setValue(float(start))
        self.params_table.setCellWidget(row, 1, start_box)

        stop_box = QDoubleSpinBox()
        stop_box.setRange(-1e9, 1e9)
        stop_box.setDecimals(2)
        stop_box.setValue(float(stop))
        self.params_table.setCellWidget(row, 2, stop_box)

        steps_box = QSpinBox()
        steps_box.setRange(1, 1000)
        steps_box.setValue(int(steps))
        self.params_table.setCellWidget(row, 3, steps_box)

        kind_combo = QComboBox()
        kind_combo.addItems(["int", "float"])
        kind_combo.setCurrentText(kind)
        self.params_table.setCellWidget(row, 4, kind_combo)

    # UI-Action: fügt eine neue Param-Zeile hinzu.
    def add_param_row(self) -> None:
        self._add_param_row_default("MUTATION_ROT_PROB", 0.05, 0.30, 4, "float")

    # UI-Action: löscht die aktuell selektierte Param-Zeile.
    def del_param_row(self) -> None:
        row = self.params_table.currentRow()
        if row >= 0:
            self.params_table.removeRow(row)

    # Liest alle Param-Zeilen aus der Tabelle in DOEParam-Strukturen.
    def _read_params(self) -> List[DOEParam]:
        params: List[DOEParam] = []
        for r in range(self.params_table.rowCount()):
            key = self.params_table.cellWidget(r, 0).currentText().strip()
            start = float(self.params_table.cellWidget(r, 1).value())
            stop = float(self.params_table.cellWidget(r, 2).value())
            steps = int(self.params_table.cellWidget(r, 3).value())
            kind = self.params_table.cellWidget(r, 4).currentText().strip()
            params.append(DOEParam(key=key, start=start, stop=stop, steps=steps, kind=kind))
        return params

    # Startet DOE: erzeugt Designpunkte, startet Worker-Thread und bereitet Ergebnis-Tabelle vor.
    def run_doe(self) -> None:
        if not self._ensure_layout_loaded():
            return

        params = self._read_params()
        if not params:
            QMessageBox.information(self, "DOE", "Bitte mindestens 1 Parameter hinzufügen.")
            return

        points = _design_points(params)
        if not points:
            QMessageBox.information(self, "DOE", "Keine Design-Punkte erzeugt.")
            return

        # Results table columns = Run + BestScore + param keys
        keys = [p.key for p in params]
        self.results_table.clear()
        self.results_table.setRowCount(0)
        self.results_table.setColumnCount(2 + len(keys))
        self.results_table.setHorizontalHeaderLabels(["Durchlauf", "Bester score"] + keys)
        self.results_table.horizontalHeader().setStretchLastSection(True)
        self._best_by_run.clear()

        gens = int(config.GENERATIONS)

        self.btn_run.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.status.setText(f"DOE startet: {len(points)} Runs")
        QApplication.processEvents()

        # Thread
        self._worker_thread = QThread()
        self._worker = DOEWorker(generations=gens, points=points)
        self._worker.moveToThread(self._worker_thread)

        self._worker_thread.started.connect(self._worker.run)
        self._worker.progressed.connect(self._on_progress)
        self._worker.result_ready.connect(self._on_result)
        self._worker.finished.connect(self._on_finished)
        self._worker.finished.connect(self._worker_thread.quit)
        self._worker_thread.finished.connect(self._worker_thread.deleteLater)
        self._worker_thread.finished.connect(self._on_thread_finished)
        self._worker_thread.finished.connect(self._worker.deleteLater)

        self._worker_thread.start()

    # UI-Action: fordert Stop an (setzt Worker-Flag und config.STOP_REQUESTED)
    def stop_doe(self) -> None:
        if self._worker:
            self.status.setText("Stop angefordert...")
            self._worker.request_stop()

    # Callback: aktualisiert Status-Text während DOE-Runs laufen
    def _on_progress(self, i: int, total: int, msg: str) -> None:
        self.status.setText(f"{msg} ({i}/{total})")
        QApplication.processEvents()

    # Callback: übernimmt Ergebnis eines Runs in die Tabelle und speichert best_ind
    def _on_result(self, run_idx: int, point: dict, score: float, best_ind: object) -> None:
        row = self.results_table.rowCount()
        self.results_table.insertRow(row)
        
        hdr = self.results_table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        for c in range(2, self.results_table.columnCount()):
            hdr.setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)
        self.results_table.setItem(row, 0, QTableWidgetItem(str(run_idx)))
        self.results_table.setItem(row, 1, QTableWidgetItem(f"{score:.4f}"))

        # param columns
        for c, k in enumerate(
            [self.results_table.horizontalHeaderItem(i).text() for i in range(2, self.results_table.columnCount())],
            start=2,
        ):
            v = point.get(k)
            self.results_table.setItem(row, c, QTableWidgetItem(str(v)))

        self._best_by_run[run_idx] = best_ind

    # Callback: räumt UI/Worker-State nach DOE-Ende auf
    def _on_finished(self) -> None:
        self.btn_run.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.status.setText("DOE fertig")
    def _on_thread_finished(self) -> None:
        self.status.setText("DOE fertig")
        self._worker = None
        self._worker_thread = None

    # Öffnet den BestDialog für den selektierten DOE-Run (wenn vorhanden)
    def view_selected_best(self) -> None:
        row = self.results_table.currentRow()
        if row < 0:
            QMessageBox.information(self, "View", "Bitte eine Ergebnis-Zeile auswählen.")
            return
        run_idx_item = self.results_table.item(row, 0)
        if not run_idx_item:
            return
        run_idx = int(run_idx_item.text())
        best_ind = self._best_by_run.get(run_idx)
        if not best_ind:
            QMessageBox.information(self, "View", "Kein BestLayout gespeichert (evtl. Run fehlgeschlagen).")
            return
        BestDialog(best_ind, parent=self, title=f"Bestes Layout (Run {run_idx})").exec()

    # Stoppt Worker/Thread sauber, damit Qt beim Beenden keinen laufenden Thread zerstört.
    def _shutdown_worker(self) -> None:
        if not self._worker_thread:
            return
        if not self._worker_thread.isRunning():
            return

        # Stop anfordern (GA soll abbrechen)
        if self._worker:
            self._worker.request_stop()

        # Warten: run_ga blockiert, daher aktiv warten und Events pumpen
        for _ in range(200):  # ~10s
            if not self._worker_thread.isRunning():
                break
            QCoreApplication.processEvents()
            self._worker_thread.wait(50)

        # Notbremse: wenn GA nicht reagiert, Thread hart beenden
        if self._worker_thread.isRunning():
            self._worker_thread.terminate()
            self._worker_thread.wait(2000)

    # Beim Fenster-Schließen sicherstellen, dass der Thread beendet ist.
    def closeEvent(self, event) -> None:
        self._shutdown_worker()
        super().closeEvent(event)

#Startet die Qt-App und zeigt das DOE-Fenster.
def main() -> int:
    app = QApplication(sys.argv)
    w = DOERunnerWindow()
    w.show()

    #Shutdown-Hooks (decken auch IDE-Stop / Exit / aboutToQuit ab)
    app.aboutToQuit.connect(w._shutdown_worker)
    atexit.register(w._shutdown_worker)

    #Exceptions: erst Worker stoppen, dann Standard-Handler
    old_hook = sys.excepthook

    def _hook(exc_type, exc, tb):
        try:
            w._shutdown_worker()
        finally:
            old_hook(exc_type, exc, tb)

    sys.excepthook = _hook

    return app.exec()



if __name__ == "__main__":
    raise SystemExit(main())