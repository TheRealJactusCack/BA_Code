"""
Kompakter PyQt6 Layout-Viewer (Wände + Säulen) — vereinfachte, lesbare Version.
Benötigt: `pyqt6`, `openpyxl`.
Aufruf: `python layout_viewer_pyqt6.py layout.xlsx`
"""
from __future__ import annotations

import math
import sys
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from PyQt6.QtCore import Qt, QPointF, QRectF
from PyQt6.QtGui import QBrush, QPen, QColor, QPainter
from PyQt6.QtWidgets import (
    QApplication, QComboBox, QFileDialog, QGraphicsEllipseItem, QGraphicsLineItem,
    QGraphicsRectItem, QGraphicsScene, QGraphicsView, QHBoxLayout, QLabel,
    QMessageBox, QPushButton, QVBoxLayout, QWidget,
)


@dataclass
class Wall:
    x1: float; 
    y1: float; 
    x2: float; 
    y2: float


@dataclass
class Column:
    x: float; 
    y: float; 
    w: Optional[float] = None; 
    d: Optional[float] = None; 
    rot: float = 0.0


@dataclass
class Machine:
    id: str
    w: Optional[float] = None
    d: Optional[float] = None
    input: float = 0.0


def _f(v: Any) -> Optional[float]:
    try:
        return None if v is None else float(v)
    except Exception:
        return None


def read_layout_from_sheet(xlsx_path: str, sheet_name: str) -> Tuple[List[Wall], List[Column], List[Machine], Dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]
    header_row = None

    for r in range(1, 31):
        a = str(ws.cell(r, 1).value or "").strip().lower()
        b = str(ws.cell(r, 2).value or "").strip().lower()
        if a == "id" and b == "type":
            header_row = r; break
    if header_row is None:
        raise ValueError("Header (id,type) nicht gefunden")

    header_map: Dict[str, int] = {}
    for c in range(1, 60):
        val = ws.cell(header_row, c).value
        if val is None: break
        header_map[str(val).strip().lower()] = c

    def idx(name: str) -> Optional[int]:
        return header_map.get(name)

    walls: List[Wall] = []
    cols: List[Column] = []
    machines: List[Machine] = []
    meta: Dict[str, Any] = {}

    for r in range(1, header_row):
        k = ws.cell(r, 1).value
        if k is None: continue
        meta[str(k).strip().lower()] = ws.cell(r, 2).value

    blank = 0
    for r in range(header_row + 1, ws.max_row + 1):
        raw_id = ws.cell(r, idx('id') or 0).value
        if raw_id is None or str(raw_id).strip() == "":
            blank += 1
            if blank >= 5: break
            continue
        blank = 0
        t = str(ws.cell(r, idx('type') or 0).value or "").strip().lower()
        if t == 'wall':
            x1 = _f(ws.cell(r, idx('x1') or 0).value); y1 = _f(ws.cell(r, idx('y1') or 0).value)
            x2 = _f(ws.cell(r, idx('x2') or 0).value); y2 = _f(ws.cell(r, idx('y2') or 0).value)
            if None not in (x1, y1, x2, y2): walls.append(Wall(x1, y1, x2, y2))
        elif t == 'column':
            x = _f(ws.cell(r, idx('x') or 0).value)
            y = _f(ws.cell(r, idx('y') or 0).value)
            if x is None or y is None: continue
            w = _f(ws.cell(r, idx('w') or 0).value)
            d = _f(ws.cell(r, idx('d') or 0).value)
            rot = _f(ws.cell(r, idx('rot') or 0).value) or 0.0
            cols.append(Column(x=x, y=y, w=w, d=d, rot=rot))
        elif t == 'machine':
            # machines: Größe w und d, Input = Flussrichtung
            label = str(raw_id)
            wv = _f(ws.cell(r, idx('w') or 0).value)
            dv = _f(ws.cell(r, idx('d') or 0).value)
            inputv = _f(ws.cell(r, idx('input') or 0).value) or 0.0
            machines.append(Machine(id=label, w=wv, d=dv, input=inputv))
        else:
            continue
    return walls, cols, machines, meta


class ZoomPanView(QGraphicsView):
    def __init__(self, scene=None):
        super().__init__(scene)
        self.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        self._panning = False; self._pan_start = None
        self.setDragMode(QGraphicsView.DragMode.NoDrag)

    def wheelEvent(self, ev):
        f = 1.15 if ev.angleDelta().y() > 0 else 1/1.15; self.scale(f, f)

    def mousePressEvent(self, ev):
        if ev.button() == Qt.MouseButton.MiddleButton:
            self._panning = True; self._pan_start = ev.pos(); self.setCursor(Qt.CursorShape.ClosedHandCursor); ev.accept(); return
        super().mousePressEvent(ev)

    def mouseMoveEvent(self, ev):
        if self._panning and self._pan_start is not None:
            d = ev.pos() - self._pan_start; self._pan_start = ev.pos()
            self.horizontalScrollBar().setValue(self.horizontalScrollBar().value() - d.x())
            self.verticalScrollBar().setValue(self.verticalScrollBar().value() - d.y()); ev.accept(); return
        super().mouseMoveEvent(ev)

    def mouseReleaseEvent(self, ev):
        if ev.button() == Qt.MouseButton.MiddleButton:
            self._panning = False; self._pan_start = None; self.setCursor(Qt.CursorShape.ArrowCursor); ev.accept(); return
        super().mouseReleaseEvent(ev)


class LayoutViewer(QWidget):
    def __init__(self, xlsx_path: Optional[str] = None):
        super().__init__(); self.setWindowTitle("Layout Viewer")
        self.xlsx_path = None; self.sheet_names: List[str] = []
        self.file_label = QLabel("No file loaded")
        self.sheet_combo = QComboBox(); self.sheet_combo.currentTextChanged.connect(self._redraw)
        self.open_btn = QPushButton("Open Excel…"); self.open_btn.clicked.connect(self._open_file)
        top = QHBoxLayout(); top.addWidget(self.open_btn); top.addWidget(QLabel("Layout:")); top.addWidget(self.sheet_combo,1); top.addWidget(self.file_label,2)
        self.scene = QGraphicsScene(self); self.view = ZoomPanView(self.scene); self.view.setBackgroundBrush(QBrush(QColor("#FAFAFA")))
        self.meta_label = QLabel(""); self.meta_label.setStyleSheet("color:#444;")
        root = QVBoxLayout(self); root.addLayout(top); root.addWidget(self.view,1); root.addWidget(self.meta_label)
        if xlsx_path: self.load_file(xlsx_path)

    def _open_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "Open Excel file", "", "Excel files (*.xlsx *.xlsm *.xls)")
        if path: self.load_file(path)

    def load_file(self, path: str):
        try: wb = load_workbook(path, data_only=True); sheets = wb.sheetnames
        except Exception as e: QMessageBox.critical(self, "Error", f"Could not open workbook:\n{e}"); return
        self.xlsx_path = path; self.sheet_names = sheets
        self.sheet_combo.blockSignals(True); self.sheet_combo.clear();
        for s in sheets: self.sheet_combo.addItem(s)
        self.sheet_combo.blockSignals(False); self.file_label.setText(path); self._redraw(self.sheet_combo.currentText())

    def _redraw(self, sheet_name: str):
        if not self.xlsx_path or not sheet_name: return
        try: walls, cols, machines, meta = read_layout_from_sheet(self.xlsx_path, sheet_name)
        except Exception as e: self.scene.clear(); self.meta_label.setText(""); QMessageBox.warning(self, "Parse error", str(e)); return
        self.scene.clear()
        # compute bounds
        xs = [v for w in walls for v in (w.x1, w.x2)] + [x for c in cols for x in (c.x - (c.d/2 if c.d else (c.w or 0.5)/2), c.x + (c.d/2 if c.d else (c.w or 0.5)/2))]
        ys = [v for w in walls for v in (w.y1, w.y2)] + [y for c in cols for y in (c.y - (c.d/2 if c.d else (c.d or 0.5)/2), c.y + (c.d/2 if c.d else (c.d or 0.5)/2))]
        if not xs or not ys: 
            return
        minx, maxx, miny, maxy = min(xs), max(xs), min(ys), max(ys)
        vw = max(self.view.viewport().width(), 800); 
        vh = max(self.view.viewport().height(), 600); 
        margin = 40.0
        scale = min((vw-2*margin)/(maxx-minx or 1e-6), (vh-2*margin)/(maxy-miny or 1e-6))
        if not math.isfinite(scale) or scale <= 0: 
            scale = 60.0

        def map_pt(x,y):
            return QPointF(margin + (x-minx)*scale, margin + (maxy-y)*scale)
        
        wall_pen = QPen(QColor("#111")); wall_pen.setWidth(2)
        col_pen = QPen(QColor("#333")); col_pen.setWidth(2); col_brush = QBrush(QColor("#C9CED6"))
        maschine_pen = QPen(QColor("#AA0000")); maschine_pen.setWidth(2); maschine_brush = QBrush(QColor("#DD8888"))


        for w in walls:
            p1, p2 = map_pt(w.x1, w.y1), map_pt(w.x2, w.y2)
            item = QGraphicsLineItem(p1.x(), p1.y(), p2.x(), p2.y()); item.setPen(wall_pen); self.scene.addItem(item)

        for c in cols:
            ctr = map_pt(c.x, c.y)
            ww = c.w if c.w is not None else 0.5; dd = c.d if c.d is not None else 0.5
            wpx, hpx = ww*scale, dd*scale
            rect = QGraphicsRectItem(ctr.x()-wpx/2, ctr.y()-hpx/2, wpx, hpx); rect.setPen(col_pen); rect.setBrush(col_brush)
            print(f"  Column at ({c.x},{c.y}): size={ww} x {dd} m")                
            if abs(c.rot) > 1e-6: rect.setTransformOriginPoint(ctr); rect.setRotation(-c.rot)
            self.scene.addItem(rect)
        frame = QGraphicsRectItem(QRectF(map_pt(minx, maxy), map_pt(maxx, miny)).normalized())
        frame_pen = QPen(QColor("#888")); frame_pen.setStyle(Qt.PenStyle.DashLine); frame_pen.setWidth(1); frame.setPen(frame_pen); frame.setBrush(QBrush(Qt.BrushStyle.NoBrush)); self.scene.addItem(frame)

        # Draw machines in a horizontal row below the layout
        if machines:
            span_x = maxx - minx
            span_y = maxy - miny
            gap = max(0.05 * span_x, 0.5)
            total_width = span_x - 2*gap
            n = len(machines)
            if n > 0:
                step = total_width / max(1, n)
                base_y_world = miny - max(0.08 * span_y, 0.5)
        # distribute centers evenly
                print("len(machines)=", len(machines) if machines else 0, flush=True)
                for i, m in enumerate(machines):
                    print(m.id, m.w, m.d, m.input)
                    cx_world = minx + gap + (i + 0.5) * step
                    cy_world = base_y_world
                    ctr = map_pt(cx_world, cy_world)
                    if m.w is None or m.d is None:
                        rpx = 0.25 * scale
                        e = QGraphicsRectItem(ctr.x()-rpx, ctr.y()-rpx, rpx, rpx)
                        e.setPen(maschine_pen); e.setBrush(maschine_brush); self.scene.addItem(e)
                        print(f"  Machine {m.id}: default size square")
                    else:
                        wpx, hpx = m.w * scale, m.d * scale
                        rpx = max(wpx, hpx) / 2
                        e = QGraphicsRectItem(ctr.x()-rpx, ctr.y()-rpx, wpx, hpx); 
                        e.setPen(maschine_pen);e.setBrush(maschine_brush); 
                        self.scene.addItem(e)
                        print(f"  Machine {m.id}: size {m.w} x {m.d} m")
        self.scene.setSceneRect(0,0,vw,vh); 
        self.view.resetTransform(); 
        self.view.fitInView(self.scene.itemsBoundingRect().adjusted(-20,-20,20,20), Qt.AspectRatioMode.KeepAspectRatio)
        unit = meta.get("unit",""); 
        origin = meta.get("origin",""); 
        self.meta_label.setText(f"unit={unit} | origin={origin} | walls={len(walls)} | columns={len(cols)}")


def main():
    app = QApplication(sys.argv); path = sys.argv[1] if len(sys.argv)>1 else None
    w = LayoutViewer(xlsx_path=path); w.resize(1100,800); w.show(); sys.exit(app.exec())


if __name__ == "__main__":
    main()
