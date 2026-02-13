# FILE: doe_runner.py
"""
DOE Runner (headless) für den GA

Start:
  python doe_runner.py

Was passiert:
  - Lädt automatisch layouts_with_machines.xlsx (gleiches Verzeichnis wie diese Datei)
  - Nutzt automatisch das Sheet "Ideal (2)"
  - Führt DOE (Full-Factorial) über hart kodierte Parameterbereiche aus
  - Zeigt eine Terminal-Progressbar für alle DOE-Runs
  - Schreibt Ergebnisse in Auswertung.xlsx (gleiches Verzeichnis) in ein neues Sheet pro Start
"""

from __future__ import annotations

import itertools
import math
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Sequence

from openpyxl import Workbook, load_workbook

import config
from excel_import import apply_excel_layout_to_config
from ga_engine import run_ga


# ------------------------------ DOE model --------------------------------------


@dataclass(frozen=True)
class DOEParam:
    # Beschreibt einen DOE-Parameter als Bereich (Start/Stop) mit einer Anzahl an Steps
    key: str
    start: float
    stop: float
    steps: int
    kind: str  # "int" or "float"


# Erzeugt n gleichmäßig verteilte Werte zwischen a und b (inkl. a und b)
def _linspace(a: float, b: float, n: int) -> List[float]:
    # Sonderfälle (n<=1 oder a==b) abfangen
    if n <= 1:
        return [float(a)]
    if math.isclose(a, b):
        return [float(a)] * n
    step = (b - a) / (n - 1)
    return [a + i * step for i in range(n)]


# Castet DOE-Werte abhängig vom Parametertyp
def _cast_value(v: float, kind: str) -> Any:
    # Int-Parameter sauber runden, Float unverändert übernehmen
    if kind == "int":
        return int(round(v))
    return float(v)


# Baut alle DOE-Designpunkte (Full-Factorial)
def _design_points(params: Sequence[DOEParam]) -> List[Dict[str, Any]]:
    # Pro Parameter eine Werteliste bauen und dann kartesisches Produkt bilden
    grids: List[List[Any]] = []
    keys: List[str] = []
    for p in params:
        keys.append(p.key)
        grids.append([_cast_value(v, p.kind) for v in _linspace(p.start, p.stop, p.steps)])

    points: List[Dict[str, Any]] = []
    for combo in itertools.product(*grids):
        points.append({k: combo[i] for i, k in enumerate(keys)})
    return points


# ------------------------------ Config helpers ---------------------------------


def _set_config_value(key: str, value: Any) -> None:
    # Setzt config.<key> (mit Fallback SWAP_PROP -> SWAP_PROB) ohne config.py zu ändern
    if hasattr(config, key):
        setattr(config, key, value)
        return

    # Kompatibilität: Falls Projekt statt SWAP_PROP nur SWAP_PROB nutzt
    if key == "SWAP_PROP" and hasattr(config, "SWAP_PROB"):
        setattr(config, "SWAP_PROB", value)
        return

    raise AttributeError(f"config hat kein Attribut '{key}' (und kein kompatibles Fallback)")


def _get_config_value(key: str) -> Any:
    # Liest config.<key> (mit Fallback SWAP_PROP -> SWAP_PROB)
    if hasattr(config, key):
        return getattr(config, key)
    if key == "SWAP_PROP" and hasattr(config, "SWAP_PROB"):
        return getattr(config, "SWAP_PROB")
    return None


# ------------------------------ Terminal progress ------------------------------


def _fmt_hhmmss(seconds: float) -> str:
    # Formatiert Sekunden als HH:MM:SS
    seconds = max(0.0, float(seconds))
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = int(seconds % 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def _print_progress(done: int, total: int, started_at: float) -> None:
    # Druckt eine einfache Trackbar inkl. ETA (ohne externe Dependencies)
    width = 32
    total = max(1, int(total))
    done = min(max(0, int(done)), total)

    elapsed = time.time() - started_at
    per_item = (elapsed / done) if done > 0 else 0.0
    eta = per_item * (total - done) if done > 0 else 0.0

    filled = int(width * (done / total))
    bar = "█" * filled + " " * (width - filled)
    pct = (done / total) * 100.0

    msg = f"[{bar}] {done}/{total} ({pct:5.1f}%) | elapsed {_fmt_hhmmss(elapsed)} | ETA {_fmt_hhmmss(eta)}"
    sys.stdout.write("\r" + msg)
    sys.stdout.flush()

    # Bei Abschluss eine neue Zeile ausgeben
    if done >= total:
        sys.stdout.write("\n")
        sys.stdout.flush()


# ------------------------------ Excel output -----------------------------------


def _unique_sheet_name(wb, base: str) -> str:
    # Erzeugt einen eindeutigen Sheet-Namen (Excel limit: 31 chars)
    name = base[:31]
    if name not in wb.sheetnames:
        return name
    i = 2
    while True:
        candidate = f"{name[:28]}_{i}"
        if candidate not in wb.sheetnames:
            return candidate
        i += 1


def _append_results_to_workbook(out_path: Path, headers: List[str], rows: List[List[Any]]) -> Path:
    # Öffnet/erstellt Auswertung.xlsx und schreibt Ergebnisse in ein neues Sheet
    if out_path.exists():
        wb = load_workbook(out_path)
    else:
        wb = Workbook()
        # Default-Sheet entfernen, damit nur unsere Sheets drin sind
        if wb.active and len(wb.sheetnames) == 1:
            wb.remove(wb.active)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    sheet_name = _unique_sheet_name(wb, f"DOE_{ts}")
    ws = wb.create_sheet(sheet_name)

    ws.append(headers)
    for r in rows:
        ws.append(r)

    wb.save(out_path)
    return out_path


# ------------------------------ Main runner ------------------------------------


def main() -> int:
    # Pfade relativ zur Datei auflösen (VM/headless)
    here = Path(__file__).resolve().parent
    layout_xlsx = here / "layouts_with_machines.xlsx"
    out_xlsx = here / "Auswertung.xlsx"
    sheet_name = "Ideal (2)"

    # Layout aus Excel in config übernehmen
    apply_excel_layout_to_config(str(layout_xlsx), sheet_name=sheet_name)

    # DOE-Parameter hier festlegen (statt UI)
    params: List[DOEParam] = [
        DOEParam(key="POPULATION_SIZE", start=10, stop=200, steps=4, kind="int"),
        DOEParam(key="GENERATIONS", start=50, stop=1000, steps=4, kind="int"),
        DOEParam(key="SWAP_PROP", start=0.00, stop=0.20, steps=3, kind="float"),
    ]

    # Alle Designpunkte erzeugen (Full-Factorial)
    points = _design_points(params)
    if not points:
        raise RuntimeError("Keine Design-Punkte erzeugt (params leer?)")

    # Ergebnisse sammeln (für Excel-Export)
    headers = ["Run", "BestScore"] + [p.key for p in params]
    rows: List[List[Any]] = []

    # Progressbar initialisieren
    started_at = time.time()
    total_runs = len(points)
    _print_progress(0, total_runs, started_at)

    # DOE Runs sequentiell ausführen (ohne UI)
    for run_idx, point in enumerate(points, start=1):
        # Pro Run DOE-Parameter in config setzen
        for k, v in point.items():
            _set_config_value(k, v)

        # Pro Run Stop-Flag zurücksetzen
        config.STOP_REQUESTED = False

        # GA laufen lassen (Generationen je Run aus config lesen)
        generations = int(_get_config_value("GENERATIONS") or config.GENERATIONS)
        best_ind, best_score = run_ga(generations, progress_callback=None)

        # Score robust behandeln, falls GA kein Ergebnis liefert
        score = float(best_score) if best_ind is not None else float("inf")

        # Ergebniszeile bauen
        row = [run_idx, score] + [point.get(p.key) for p in params]
        rows.append(row)

        # Progressbar updaten
        _print_progress(run_idx, total_runs, started_at)

    # Ergebnisse in Auswertung.xlsx in ein neues Sheet schreiben
    _append_results_to_workbook(out_xlsx, headers=headers, rows=rows)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
