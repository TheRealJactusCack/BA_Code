# UI MAIN MODULE

from __future__ import annotations

from PyQt6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QGroupBox,
    QLabel,
    QPushButton,
    QSpinBox,
    QDoubleSpinBox,
    QMessageBox,
    QApplication,
    QFileDialog,
    QComboBox,
    QGridLayout,
)

import config
from helpers import update_grid_counts
from ga_engine import run_ga
from ui_canvas import PopulationCanvas, BestDialog
from excel_import import apply_excel_layout_to_config


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Fabrikplaner GA")
        self.setMinimumSize(1024, 768)

        self.layout_loaded = False
        self.xlsx_path: str | None = None

        root = QVBoxLayout(self)
        controls = QHBoxLayout()
        root.addLayout(controls)

        layout_group = QGroupBox("Layout Excel")
        controls.addWidget(layout_group)
        layout_v = QVBoxLayout(layout_group)

        self.excel_file_label = QLabel("Keine Excel Datei geladen")
        self.sheet_combo = QComboBox()
        self.sheet_combo.setEnabled(False)
        self.sheet_combo.currentTextChanged.connect(self._on_sheet_changed)

        self.load_excel_btn = QPushButton("Excel laden")
        self.load_excel_btn.clicked.connect(self.load_excel)

        self.reload_excel_btn = QPushButton("Layout neu laden")
        self.reload_excel_btn.clicked.connect(self.reload_excel)
        self.reload_excel_btn.setEnabled(False)

        self.machine_info_label = QLabel(f"Maschinen {config.MACHINE_COUNT}")

        layout_v.addWidget(self.load_excel_btn)
        layout_v.addWidget(self.reload_excel_btn)
        layout_v.addWidget(QLabel("Sheet"))
        layout_v.addWidget(self.sheet_combo)
        layout_v.addWidget(self.excel_file_label)
        layout_v.addWidget(self.machine_info_label)


        ga_group = QGroupBox("Genetischer Algorithmus")
        controls.addWidget(ga_group)
        ga_v = QVBoxLayout(ga_group)

        self.population_size_spin = QSpinBox()
        self.population_size_spin.setRange(1, 1000)
        self.population_size_spin.setValue(config.POPULATION_SIZE)
        ga_v.addWidget(QLabel("Population Größe"))
        ga_v.addWidget(self.population_size_spin)

        self.elite_keep_spin = QSpinBox()
        self.elite_keep_spin.setRange(1, 1000)
        self.elite_keep_spin.setValue(config.ELITE_KEEP)
        ga_v.addWidget(QLabel("Eliten behalten"))
        ga_v.addWidget(self.elite_keep_spin)

        self.mutation_prob_spin = QDoubleSpinBox()
        self.mutation_prob_spin.setRange(0.0, 1.0)
        self.mutation_prob_spin.setValue(config.BASE_MUTATION_PROB)
        self.mutation_prob_spin.setSingleStep(0.01)
        ga_v.addWidget(QLabel("Mutation Position Wahrscheinlichkeit"))
        ga_v.addWidget(self.mutation_prob_spin)

        self.mutation_pos_std_spin = QDoubleSpinBox()
        self.mutation_pos_std_spin.setRange(0.1, 10.0)
        self.mutation_pos_std_spin.setValue(config.BASE_MUTATION_POS_STD)
        self.mutation_pos_std_spin.setSingleStep(0.1)
        ga_v.addWidget(QLabel("Position Mutation Std"))
        ga_v.addWidget(self.mutation_pos_std_spin)

        self.mutation_rot_prob_spin = QDoubleSpinBox()
        self.mutation_rot_prob_spin.setRange(0.0, 1.0)
        self.mutation_rot_prob_spin.setValue(config.BASE_MUTATION_ROT_PROB)
        self.mutation_rot_prob_spin.setSingleStep(0.01)
        ga_v.addWidget(QLabel("Mutation Rotation Wahrscheinlichkeit"))
        ga_v.addWidget(self.mutation_rot_prob_spin)

        self.generations_spin = QSpinBox()
        self.generations_spin.setRange(1, 10000)
        self.generations_spin.setValue(200)
        ga_v.addWidget(QLabel("Generationen"))
        ga_v.addWidget(self.generations_spin)

        self.start_btn = QPushButton("GA starten")
        self.start_btn.clicked.connect(self.start_ga)
        ga_v.addWidget(self.start_btn)

        self.stop_btn = QPushButton("Abbrechen")
        self.stop_btn.setEnabled(False)
        ga_v.addWidget(self.stop_btn)

        self.gencounter_label = QLabel("0 0")
        ga_v.addWidget(self.gencounter_label)

        self.pop_canvas = PopulationCanvas(self, cols=10, rows=5)
        root.addWidget(self.pop_canvas)

        self.status_label = QLabel("Bitte Excel Layout laden")
        root.addWidget(self.status_label)

        update_grid_counts()

    def load_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Excel Layout laden", "", "Excel files (*.xlsx *.xlsm *.xls)")
        if not path:
            return
        self.xlsx_path = path
        self._populate_sheets(path)

    def _populate_sheets(self, path: str):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            sheets = list(wb.sheetnames)
        except Exception as e:
            QMessageBox.critical(self, "Excel Fehler", f"Konnte Excel nicht öffnen\n{e}")
            return

        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        for s in sheets:
            self.sheet_combo.addItem(s)
        self.sheet_combo.blockSignals(False)

        self.sheet_combo.setEnabled(True)
        self.reload_excel_btn.setEnabled(True)
        self.excel_file_label.setText(path)

        if sheets:
            self._apply_excel(path, sheets[0])

    def reload_excel(self):
        if self.xlsx_path and self.sheet_combo.currentText():
            self._apply_excel(self.xlsx_path, self.sheet_combo.currentText())

    def _on_sheet_changed(self, sheet: str):
        if self.xlsx_path and sheet:
            self._apply_excel(self.xlsx_path, sheet)

    def _apply_excel(self, path: str, sheet: str):
        try:
            apply_excel_layout_to_config(path, sheet_name=sheet)
            self.layout_loaded = True
        except Exception as e:
            self.layout_loaded = False
            QMessageBox.warning(self, "Excel Parse Fehler", str(e))
            return

        self.machine_info_label.setText(f"Maschinen {config.MACHINE_COUNT}")
        self.status_label.setText(
            f"Layout geladen {sheet} Floor {config.FLOOR_W:.2f} x {config.FLOOR_H:.2f} Obstacles {len(config.OBSTACLES)} Flows {len(config.MATERIAL_CONNECTIONS)}"
        )

        self.pop_canvas.update()
        QApplication.processEvents()

    def _ensure_layout_loaded(self) -> bool:
        if self.layout_loaded:
            return True
        QMessageBox.information(self, "Layout fehlt", "Bitte zuerst ein Excel Layout laden")
        return False

    def start_ga(self):
        if not self._ensure_layout_loaded():
            return

        config.POPULATION_SIZE = int(self.population_size_spin.value())
        config.ELITE_KEEP = int(self.elite_keep_spin.value())
        config.MUTATION_PROB = float(self.mutation_prob_spin.value())
        config.MUTATION_POS_STD = float(self.mutation_pos_std_spin.value())
        config.MUTATION_ROT_PROB = float(self.mutation_rot_prob_spin.value())

        gens = int(self.generations_spin.value())

        self.start_btn.setEnabled(False)
        self.status_label.setText("GA läuft")
        QApplication.processEvents()

        config.STOP_REQUESTED = False
        self.stop_btn.setEnabled(True)
        try:
            self.stop_btn.clicked.disconnect()
        except Exception:
            pass
        self.stop_btn.clicked.connect(self._stop)

        def _cb(g, total, best_score_cb, best_ind_cb, population=None):
            self.gencounter_label.setText(f"{g} {total}")
            if population:
                self.pop_canvas.set_population(population)
            self.status_label.setText(f"Gen {g} bester Score {best_score_cb:.2f}")
            QApplication.processEvents()

        best_ind, best_score = run_ga(gens, progress_callback=_cb)
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)

        if best_ind:
            BestDialog(best_ind, parent=self, title="Beste Lösung").exec()
        else:
            QMessageBox.information(self, "GA beendet", f"Kein Ergebnis Best score {best_score}")

    def _stop(self):
        config.STOP_REQUESTED = True
        self.status_label.setText("Stop angefordert")
        QApplication.processEvents()
