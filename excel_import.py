# EXCEL IMPORT MODULE

from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import math
from openpyxl import load_workbook

import config
from helpers import update_grid_counts, port_world_xy


@dataclass(frozen=True)
class Wall:
    x1: float
    y1: float
    x2: float
    y2: float

@dataclass(frozen=True)
class Input:
    id: str
    x: float
    y: float

@dataclass(frozen=True)
class Output:
    id: str
    x: float
    y: float
    input: float

@dataclass(frozen=True)
class Column:
    x: float
    y: float
    w: float
    d: float
    rot: float = 0.0

@dataclass(frozen=True)
class MachineRow:
    id: str
    x: float
    y: float
    rot: float
    w: float
    d: float
    input: str
    side_in: str
    offset_in: float
    side_out: str
    offset_out: float
    worker: Optional[int]
    side_worker: Optional[str]
    offset_worker: Optional[float]
    side_water: Optional[str]
    offset_water: Optional[float]
    side_gas: Optional[str]
    offset_gas: Optional[float]
    side_other: Optional[str]
    offset_other: Optional[float]

@dataclass(frozen=True)
class Water:
    id: str
    x: float
    y: float
    input: str

@dataclass(frozen=True)
class Gas:
    id: str
    x: float
    y: float
    input: str

@dataclass(frozen=True)
class Other:
    id: str
    x: float
    y: float
    input: str

def _f(v: Any) -> Optional[float]:
    try:
        return None if v is None else float(v)
    except (ValueError, TypeError):
        return None

def _s(v: Any) -> str:
    try:
        return "" if v is None else str(v).strip()
    except (ValueError, TypeError):
        return ""


def _truthy(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return bool(v)
    s = str(v).strip().lower()
    return s in {"1", "true", "yes", "y", "ja", "on"}

def _coerce_machine_index(v: float, n: int) -> Optional[int]:
    """
    Excel Output.input -> Maschinenindex
    Akzeptiert 1-basiert (1..n) oder 0-basiert (0..n-1)
    """
    k = int(round(float(v)))

    if 1 <= k <= n:
        k -= 1

    if 0 <= k < n:
        return k
    return None

def _compute_bounds_from_walls(walls: List[Wall]) -> Optional[Tuple[float, float, float, float]]:
    if not walls:
        return None
    xs = [p for w in walls for p in (w.x1, w.x2)]
    ys = [p for w in walls for p in (w.y1, w.y2)]
    if not xs or not ys:
        return None
    return min(xs), max(xs), min(ys), max(ys)

#=== Facility-Polygon aus Walls + Outside-Cells blocken===============================

def _pt_key(p: Tuple[float, float], ndigits: int = 6) -> Tuple[float, float]:
    return (round(float(p[0]), ndigits), round(float(p[1]), ndigits))

def _walls_to_polygon(walls_xy: List[Tuple[Tuple[float, float], Tuple[float, float]]]) -> Optional[List[Tuple[float, float]]]:
    """Konstruiert ein geschlossenes N-Eck aus Wall-Segmenten
    Erwartung: jede Ecke hat Grad 2 (einfacher geschlossener Rand)
    """
    if not walls_xy:
        return None

    adj: Dict[Tuple[float, float], List[Tuple[float, float]]] = {}
    for a, b in walls_xy:
        ka = _pt_key(a)
        kb = _pt_key(b)
        adj.setdefault(ka, []).append(kb)
        adj.setdefault(kb, []).append(ka)

    # Jede Ecke muss genau 2 Nachbarn haben (Polygon-Rand)
    if any(len(nbs) != 2 for nbs in adj.values()):
        return None

    start = min(adj.keys())
    poly: List[Tuple[float, float]] = [start]
    prev: Optional[Tuple[float, float]] = None
    cur = start

    # Walk entlang der Kanten
    for _ in range(len(adj) + 5):
        n1, n2 = adj[cur]
        nxt = n1 if prev is None or n1 != prev else n2
        if nxt == start:
            return poly  # geschlossen
        poly.append(nxt)
        prev, cur = cur, nxt

    return None


def _point_on_segment(px: float, py: float, ax: float, ay: float, bx: float, by: float, eps: float = 1e-9) -> bool:
    #Kollinearität + Bounding-Box (robust genug für Grid-Center)
    vx, vy = bx - ax, by - ay
    wx, wy = px - ax, py - ay
    cross = vx * wy - vy * wx
    if abs(cross) > eps:
        return False
    dot = wx * vx + wy * vy
    if dot < -eps:
        return False
    vv = vx * vx + vy * vy
    if dot - vv > eps:
        return False
    return True


def _point_in_polygon(px: float, py: float, poly: List[Tuple[float, float]]) -> bool:
    """
    Ray-Casting Punkte auf dem Rand zählen als 'innen'
    """
    inside = False
    n = len(poly)
    for i in range(n):
        ax, ay = poly[i]
        bx, by = poly[(i + 1) % n]

        if _point_on_segment(px, py, ax, ay, bx, by):
            return True

        # Ray cast nach +x
        cond = (ay > py) != (by > py)
        if cond:
            x_int = ax + (py - ay) * (bx - ax) / ((by - ay) if (by - ay) != 0 else 1e-12)
            if x_int >= px:
                inside = not inside
    return inside


def _outside_cells_from_walls(to_internal_xy, walls: List[Wall]) -> set[Tuple[int, int]]:
    """
    Markiert alle Grid-Zellen außerhalb des durch Walls beschriebenen Polygons als blockiert
    """
    gs = float(config.GRID_SIZE) if float(config.GRID_SIZE) > 0 else 1.0

    walls_xy: List[Tuple[Tuple[float, float], Tuple[float, float]]] = []
    for w in walls:
        a = to_internal_xy(float(w.x1), float(w.y1))
        b = to_internal_xy(float(w.x2), float(w.y2))
        walls_xy.append((a, b))

    poly = _walls_to_polygon(walls_xy)
    if not poly:
        return set()

    out: set[Tuple[int, int]] = set()
    for col in range(int(config.GRID_COLS)):
        for row in range(int(config.GRID_ROWS)):
            px = (col + 0.5) * gs
            py = (row + 0.5) * gs
            if not _point_in_polygon(px, py, poly):
                out.add((col, row))
    return out


def _point_in_rotated_rect(px: float, py: float, cx: float, cy: float, w: float, h: float, rot_deg: float) -> bool:
    a = math.radians(float(rot_deg))
    ca = math.cos(-a)
    sa = math.sin(-a)
    dx = px - cx
    dy = py - cy
    lx = dx * ca - dy * sa
    ly = dx * sa + dy * ca
    return (abs(lx) <= w / 2.0 + 1e-9) and (abs(ly) <= h / 2.0 + 1e-9)


def _cells_for_rotated_rect(cx: float, cy: float, w: float, h: float, rot_deg: float) -> set[Tuple[int, int]]:
    a = math.radians(float(rot_deg))
    dx = w / 2.0
    dy = h / 2.0
    corners = [(-dx, -dy), (dx, -dy), (dx, dy), (-dx, dy)]
    ca = math.cos(a)
    sa = math.sin(a)
    pts = []
    for x, y in corners:
        rx = x * ca - y * sa + cx
        ry = x * sa + y * ca + cy
        pts.append((rx, ry))
    minx = min(p[0] for p in pts)
    maxx = max(p[0] for p in pts)
    miny = min(p[1] for p in pts)
    maxy = max(p[1] for p in pts)

    gs = float(config.GRID_SIZE)
    c0 = max(0, int(math.floor(minx / gs)))
    c1 = min(config.GRID_COLS - 1, int(math.floor(maxx / gs)))
    r0 = max(0, int(math.floor(miny / gs)))
    r1 = min(config.GRID_ROWS - 1, int(math.floor(maxy / gs)))

    out: set[Tuple[int, int]] = set()
    for col in range(c0, c1 + 1):
        for row in range(r0, r1 + 1):
            px = (col + 0.5) * gs
            py = (row + 0.5) * gs
            if _point_in_rotated_rect(px, py, cx, cy, w, h, rot_deg):
                out.add((col, row))
    return out


def read_layout_from_sheet(
    xlsx_path: str, sheet_name: str
) -> Tuple[List[Wall], List[Column], List[MachineRow], Dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name} nicht gefunden")
    ws = wb[sheet_name]

    header_row = None
    for r in range(1, 31):
        a = str(ws.cell(r, 1).value or "").strip().lower()
        b = str(ws.cell(r, 2).value or "").strip().lower()
        if a == "id" and b == "type":
            header_row = r
            break
    if header_row is None:
        raise ValueError("Header id type nicht gefunden")

    header_map: Dict[str, int] = {}
    for c in range(1, 120):
        val = ws.cell(header_row, c).value
        if val is None:
            break
        header_map[str(val).strip().lower()] = c

    def idx(name: str) -> Optional[int]:
        return header_map.get(name.strip().lower())

    meta: Dict[str, Any] = {}
    for r in range(1, header_row):
        k = ws.cell(r, 1).value
        if k is None:
            continue
        meta[str(k).strip().lower()] = ws.cell(r, 2).value

    walls: List[Wall] = []
    inputs: List[Input] = []
    outputs: List[Output] = []
    cols: List[Column] = []
    machines: List[MachineRow] = []
    water: List[Water] = []
    gas: List[Gas] = []
    other: List[Other] = []

    #======================================================================================
    #====================Hier wird eingelesen==============================================
    #======================================================================================

    blank = 0
    for r in range(header_row + 1, ws.max_row + 1):
        raw_id = ws.cell(r, idx("id") or 0).value
        if raw_id is None or str(raw_id).strip() == "":
            blank += 1
            if blank >= 5:
                break
            continue
        blank = 0

        t = str(ws.cell(r, idx("type") or 0).value or "").strip().lower()

        if t == "wall":
            x1 = _f(ws.cell(r, idx("x1") or 0).value)
            y1 = _f(ws.cell(r, idx("y1") or 0).value)
            x2 = _f(ws.cell(r, idx("x2") or 0).value)
            y2 = _f(ws.cell(r, idx("y2") or 0).value)
            if None not in (x1, y1, x2, y2):
                walls.append(Wall(x1=x1, y1=y1, x2=x2, y2=y2))

        elif t == "input":
            iid = str(raw_id).strip()
            x = _f(ws.cell(r, idx("x") or 0).value) 
            y = _f(ws.cell(r, idx("y") or 0).value) 
            if x is None or y is None:
                raise ValueError(f"Input benötigt x und y bei Zeile {r}")
            inputs.append(Input(id=iid, x=float(x), y=float(y)))
            #print(f"Input {iid} bei ({x:.2f}, {y:.2f})")

        elif t == "output":
            oid = str(raw_id).strip()
            x = _f(ws.cell(r, idx("x") or 0).value)
            y = _f(ws.cell(r, idx("y") or 0).value)
            input_machine = _f(ws.cell(r, idx("input") or 0).value)
            if x is None or y is None:
                raise ValueError(f"Output benötigt x und y bei Zeile {r}")
            if input_machine is None:
                raise ValueError(f"Output benötigt input bei Zeile {r}")
            outputs.append(Output(id=oid, x=float(x), y=float(y), input=float(input_machine)))
            #print(f"Output {oid} bei ({x:.2f}, {y:.2f})")


        elif t == "column":
            x = _f(ws.cell(r, idx("x") or 0).value)
            y = _f(ws.cell(r, idx("y") or 0).value)
            if x is None or y is None:
                continue
            w = _f(ws.cell(r, idx("w") or 0).value) or 0.5
            d = _f(ws.cell(r, idx("d") or 0).value) or 0.5
            rot = _f(ws.cell(r, idx("rot") or 0).value) or 0.0
            cols.append(Column(x=float(x), y=float(y), w=float(w), d=float(d), rot=float(rot)))

        elif t == "machine":
            mid = str(raw_id).strip()

            #x/y werden als optional eingelesen, damit es feste, nicht bewegbare maschinen gibt
            mx = _f(ws.cell(r, idx("x")).value) if idx("x") else None
            x = None if mx is None else int(mx)
            my = _f(ws.cell(r, idx("y")).value) if idx("y") else None
            y = None if my is None else int(my)
            mr = _f(ws.cell(r, idx("rot")).value) if idx("rot") else None
            rot = None if mr is None else int(mr)

            w = _f(ws.cell(r, idx("w") or 0).value) or 1.0
            d = _f(ws.cell(r, idx("d") or 0).value) or 1.0
            input_raw = _s(ws.cell(r, idx("input") or 0).value) if idx("input") else ""

            side_in = _s(ws.cell(r, idx("side_in") or 0).value) if idx("side_in") else "Left"
            offset_in = _f(ws.cell(r, idx("offset_in") or 0).value) if idx("offset_in") else None
            side_out = _s(ws.cell(r, idx("side_out") or 0).value) if idx("side_out") else "Right"
            offset_out = _f(ws.cell(r, idx("offset_out") or 0).value) if idx("offset_out") else None

            fv = _f(ws.cell(r, idx("worker")).value) if idx("worker") else None
            worker = None if fv is None else int(fv)

            # side/offset nur dann setzen, wenn worker wirklich existiert
            if worker is None:
                side_w = None
                offset_w = None
                raw_offset_w = None
            else:
                side_w = (
                    _s(ws.cell(r, c).value)
                    if (c := idx("side_worker")) and ws.cell(r, c).value not in (None, "")
                    else None
                )
                raw_offset_w = (
                    _f(ws.cell(r, c).value)
                    if (c := idx("offset_worker")) and ws.cell(r, c).value not in (None, "")
                    else None
                )

                #Offset ist abhängig von den Kanten:
                #left/right -> entlang d, top/bottom -> entlang w
                if side_w is None:
                    offset_w = None
                elif raw_offset_w is None:
                    if str(side_w).strip().lower() in {"left", "right"}:
                        offset_w = float(d) / 2.0
                    else:
                        offset_w = float(w) / 2.0
                else:
                    offset_w = float(raw_offset_w)

            if offset_in is None:
                offset_in = float(d) / 2.0
            if offset_out is None:
                offset_out = float(d) / 2.0

            #ports für die Anschlüsse der Maschinen
            side_water = (_s(ws.cell(r, c).value) if (c := idx("water_in")) and ws.cell(r, c).value not in (None, "") else None)
            raw_offset_water = (_f(ws.cell(r, c).value) if (c := idx("offset_water")) and ws.cell(r, c).value not in (None, "") else None)

            side_gas = (_s(ws.cell(r, c).value) if (c := idx("gas_in")) and ws.cell(r, c).value not in (None, "") else None)
            raw_offset_gas = (_f(ws.cell(r, c).value) if (c := idx("offset_gas")) and ws.cell(r, c).value not in (None, "") else None)          
            
            side_other = (_s(ws.cell(r, c).value) if (c := idx("other_in")) and ws.cell(r, c).value not in (None, "") else None)
            raw_offset_other = (_f(ws.cell(r, c).value) if (c := idx("offset_other")) and ws.cell(r, c).value not in (None, "") else None)

            machines.append(
                MachineRow(
                    id = mid,
                    x = float(x) if x is not None else None,
                    y = float(y) if y is not None else None,
                    rot = float(rot) if rot is not None else None,
                    w = float(w),
                    d = float(d),
                    input = input_raw,
                    side_in = side_in,
                    offset_in = float(offset_in),
                    side_out = side_out,
                    offset_out = float(offset_out),
                    worker = worker,
                    side_worker = side_w,
                    offset_worker = float(offset_w) if offset_w is not None else None,
                    side_water = side_water,
                    offset_water = float(raw_offset_water) if raw_offset_water is not None else None,
                    side_gas = side_gas,
                    offset_gas = float(raw_offset_gas) if raw_offset_gas is not None else None,
                    side_other = side_other,
                    offset_other = float(raw_offset_other) if raw_offset_other is not None else None
                )
            )

        elif t == "water":
            wid = str(raw_id).strip()
            w_x = _f(ws.cell(r, idx("x") or 0).value)
            w_y = _f(ws.cell(r, idx("y") or 0).value)
            water_machine = _s(ws.cell(r, idx("input") or 0).value)
            if w_x is None or w_y is None:
                raise ValueError(f"Output benötigt x und y bei Zeile {r}")
            if water_machine is None:
                raise ValueError(f"Output benötigt input bei Zeile {r}")
            water.append(Water(id=wid, x=float(w_x), y=float(w_y), input=water_machine))
            #print(f"Water {wid} bei ({w_x:.2f}, {w_y:.2f})")

        elif t == "gas":
            gid = str(raw_id).strip()
            g_x = _f(ws.cell(r, idx("x") or 0).value)
            g_y = _f(ws.cell(r, idx("y") or 0).value)
            gas_machine = _s(ws.cell(r, idx("input") or 0).value)
            if g_x is None or g_y is None:
                raise ValueError(f"Output benötigt x und y bei Zeile {r}")
            if gas_machine is None:
                raise ValueError(f"Output benötigt input bei Zeile {r}")
            gas.append(Gas(id=gid, x=float(g_x), y=float(g_y), input=gas_machine))
            #print(f"Output {gid} bei ({g_x:.2f}, {g_y:.2f})")

        elif t == "other":
            oid = str(raw_id).strip()
            o_x = _f(ws.cell(r, idx("x") or 0).value)
            o_y = _f(ws.cell(r, idx("y") or 0).value)
            other_machine = _s(ws.cell(r, idx("input") or 0).value)
            if o_x is None or o_y is None:
                raise ValueError(f"Output benötigt x und y bei Zeile {r}")
            if other_machine is None:
                raise ValueError(f"Output benötigt input bei Zeile {r}")
            other.append(Other(id=oid, x=float(o_x), y=float(o_y), input=other_machine))
            #print(f"Output {oid} bei ({o_x:.2f}, {o_y:.2f})")

    return walls, inputs, outputs, cols, machines, water, gas, other, meta

    #======================================================================================
    #====================Hier wird was anders gemacht======================================
    #======================================================================================


def apply_excel_layout_to_config(xlsx_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True)
    if not wb.sheetnames:
        raise ValueError("Workbook hat keine Sheets")
    sheet = sheet_name or wb.sheetnames[0]

    walls, inputs, outputs, cols, machines, water, gas, other, meta = read_layout_from_sheet(xlsx_path, sheet)

    if "grid_size" in meta and _f(meta["grid_size"]) is not None:
        config.GRID_SIZE = float(_f(meta["grid_size"]))

    bounds = _compute_bounds_from_walls(walls)
    if bounds is not None:
        minx, maxx, miny, maxy = bounds
        floor_w = float(maxx - minx)
        floor_h = float(maxy - miny)
    else:
        minx, miny = 0.0, 0.0
        floor_w = float(_f(meta.get("floor_w")) or config.FLOOR_W)
        floor_h = float(_f(meta.get("floor_h")) or config.FLOOR_H)
        maxy = miny + floor_h

    mw = _f(meta.get("floor_w"))
    mh = _f(meta.get("floor_h"))
    if mw is not None:
        floor_w = float(mw)
    if mh is not None:
        floor_h = float(mh)

    config.FLOOR_W = max(1e-6, float(floor_w))
    config.FLOOR_H = max(1e-6, float(floor_h))

    update_grid_counts()

    flip_y = True if "flip_y" not in meta else _truthy(meta.get("flip_y"))
    maxy_for_flip = maxy if bounds is not None else (miny + floor_h)

    def to_internal_xy(x: float, y: float) -> Tuple[float, float]:
        x0 = float(x) - float(minx)
        if flip_y:
            y0 = float(maxy_for_flip) - float(y)
        else:
            y0 = float(y) - float(miny)
        return x0, y0

    if machines:
        config.MACHINE_COUNT = int(len(machines))
        config.MACHINE_LABELS = [m.id for m in machines]
        config.MACHINE_SIZES = [(max(0.1, m.w), max(0.1, m.d)) for m in machines]
        
        n = len(machines)
        # Fixed machines: aligned list (idx == machine idx)
        config.MACHINE_FIXED = [None] * n

        for i, m in enumerate(machines):
            # fixed wenn x & y gesetzt (rot optional)
            if m.x is not None and m.y is not None:
                config.MACHINE_FIXED[i] = {
                    "x": float(m.x),
                    "y": float(m.y),
                    "z": int(m.rot) if m.rot is not None else None,
                }

        # Ports sind Pflicht -> Liste n lang
        config.MACHINE_PORTS = [None] * n

        # Optionale Dinge: Liste n lang, Einträge sind None oder Dict
        config.MACHINE_WORKERS = [None] * n
        config.MACHINE_WATER = [None] * n
        config.MACHINE_GAS = [None] * n
        config.MACHINE_OTHER = [None] * n

        for i, m in enumerate(machines):
            #Pflicht-Ports immer setzen
            config.MACHINE_PORTS[i] = {
                "side_in": str(m.side_in).strip(),
                "offset_in": float(m.offset_in),
                "side_out": str(m.side_out).strip(),
                "offset_out": float(m.offset_out),
            }

            #Worker nur setzen, wenn vollständig vorhanden
            if m.worker is not None and m.side_worker is not None and m.offset_worker is not None:
                config.MACHINE_WORKERS[i] = {
                    "worker": int(m.worker),
                    "side_worker": str(m.side_worker).strip(),
                    "offset_worker": float(m.offset_worker),
                }

            #Wasser/Gas/Other nur setzen, wenn vollständig vorhanden
            if m.side_water is not None and m.offset_water is not None:
                config.MACHINE_WATER[i] = {
                    "side_water": str(m.side_water).strip(),
                    "offset_water": float(m.offset_water),
                }
            #Gas
            if m.side_gas is not None and m.offset_gas is not None:
                config.MACHINE_GAS[i] = {
                    "side_gas": str(m.side_gas).strip(),
                    "offset_gas": float(m.offset_gas),
                }
            #Other
            if m.side_other is not None and m.offset_other is not None:
                config.MACHINE_OTHER[i] = {
                    "side_other": str(m.side_other).strip(),
                    "offset_other": float(m.offset_other),
                }
            #print(f"Machine {m.id} mit Größe ({m.w:.2f}, {m.d:.2f}) und Worker {m.worker} an Seite {m.side_worker} mit Offset {m.offset_worker}")

        id_to_idx = {m.id: i for i, m in enumerate(machines)}

        #WORKER_CONNECTIONS als LOOP pro Worker (Excel-Reihenfolge)
        #Für 2 Maschinen entsteht auch A->B und B->A
        worker_to_seq: Dict[int, List[int]] = {}
        for m in machines:
            w_id = int(m.worker) if m.worker is not None else None
            if w_id is not None:
                worker_to_seq.setdefault(w_id, []).append(id_to_idx[m.id])

        w_edges: List[Tuple[Optional[int], Optional[int]]] = []
        for w_id, seq in worker_to_seq.items():
            if len(seq) < 2:
                continue
            #Nachbarn verbinden
            for i in range(len(seq) - 1):
                w_edges.append((seq[i], seq[i + 1]))
            #Loop schließen: letzter zu erster
            w_edges.append((seq[-1], seq[0]))
        config.WORKER_CONNECTIONS = w_edges

        input_obj = inputs[0] if inputs else None
        output_obj = outputs[0] if outputs else None

        input_id = str(input_obj.id).strip() if input_obj else None

        edges: List[Tuple[Optional[int], Optional[int]]] = []

        # 1) Maschinen->Maschinen + (Quelle->Maschine) explizit speichern
        for i, m in enumerate(machines):
            raw = str(m.input or "").strip()
            if not raw:
                continue
            parts = [p.strip() for p in raw.replace(";", ",").split(",") if p.strip()]
            for up in parts:
                if input_id and up == input_id:
                    edges.append((None, i))
                    continue

                if up in id_to_idx:
                    edges.append((id_to_idx[up], i))
                    continue

                raise ValueError(f"input referenziert unbekannte Maschine/Quelle {up} bei {m.id}")

        # 2) (Maschine->Senke) explizit aus Output.input speichern (wenn vorhanden)
        if output_obj is not None:
            out_target: Optional[int] = None

            raw_out = str(output_obj.input).strip() if output_obj.input is not None else ""
            raw_out_int = str(int(round(float(output_obj.input)))) if output_obj.input is not None else ""

            # a) Erst als Maschinen-ID matchen (z.B. "13")
            if raw_out in id_to_idx:
                out_target = id_to_idx[raw_out]
            elif raw_out_int in id_to_idx:
                out_target = id_to_idx[raw_out_int]
            else:
                # b) Fallback: als Index (0-basiert oder 1-basiert) interpretieren
                out_target = _coerce_machine_index(output_obj.input, len(machines))

            if out_target is not None:
                edges.append((out_target, None))
            else:
                print(f"WARN: Output.input={output_obj.input} konnte weder als ID noch als Index gemappt werden")

        # 3) Falls Entry/Exit nicht explizit vorhanden: aus mm-Kanten ableiten (Fallback)
        mm_edges: List[Tuple[int, int]] = [
            (int(a), int(b)) for (a, b) in edges if a is not None and b is not None
        ]

        incoming = [0] * len(machines)
        outgoing = [0] * len(machines)
        for a, b in mm_edges:
            if 0 <= a < len(machines) and 0 <= b < len(machines):
                outgoing[a] += 1
                incoming[b] += 1

        sources = [i for i in range(len(machines)) if incoming[i] == 0]
        sinks = [i for i in range(len(machines)) if outgoing[i] == 0]

        has_entry = any(a is None for (a, _) in edges)
        has_exit = any(b is None for (_, b) in edges)

        if input_obj is not None and not has_entry:
            for s in sources:
                edges.append((None, s))
                #print(f"Verbindung {input_id} -> {machines[s].id} (entry, auto)")

        if output_obj is not None and not has_exit:
            for t in sinks:
                edges.append((t, None))
                #print(f"Verbindung {machines[t].id} -> {output_obj.id} (exit, auto)")

        config.MATERIAL_CONNECTIONS = edges
    else:
        config.MACHINE_PORTS = []
        config.MATERIAL_CONNECTIONS = []

    if inputs and outputs:
        first_input = inputs[0]
        ix_w, iy_w = to_internal_xy(float(first_input.x), float(first_input.y))

        first_output = outputs[0]
        ox_w, oy_w = to_internal_xy(float(first_output.x), float(first_output.y))

        gs = float(config.GRID_SIZE) if float(config.GRID_SIZE) > 0 else 1.0

        # ENTRY_CELL/EXIT_CELL sind Rasterzellen (col,row), nicht Meter.
        config.ENTRY_CELL = (ix_w / gs, iy_w / gs)
        config.EXIT_CELL = (ox_w / gs, oy_w / gs)

    #======================================================================================
    #====================Hier wird Wasser, gas und other zu wegen gemacht==================
    #======================================================================================

    #print(f"Parsed {len(walls)} Walls, {len(cols)} Columns, {len(machines)} Machines, {len(inputs)} Inputs, {len(outputs)} Outputs, {len(water)} Water, {len(gas)} Gas, {len(other)} Other")
    if water:
        config.WATER_CELL = to_internal_xy(float(water[0].x), float(water[0].y))
        Water_obj = water[0] #objekt aus liste holen, da es nur eines geben soll laut excel

        water_edges: List[Tuple[Optional[int], Optional[int]]] = [] 
        raw_w = str(Water_obj.input or "").strip()
        w_parts = []
        w_parts = [p.strip() for p in raw_w.replace(";", ",").split(",") if p.strip()] #input in Liste einteilen (als string) und leere Einträge entfernen
        for i, w in enumerate(w_parts):
            if not w:
                continue
            water_edges.append((int(Water_obj.id), id_to_idx[w])) #Edges für Wasserverbindung erstellen, von Wasserquelle (Water_obj.id) zu Ziel (w) aber als int

    if gas:
        config.GAS_CELL = to_internal_xy(float(gas[0].x), float(gas[0].y))
        Gas_obj = gas[0] #objekt aus liste holen, da es nur eines geben soll laut excel

        gas_edges: List[Tuple[Optional[int], Optional[int]]] = [] 
        raw_g = str(Gas_obj.input or "").strip()
        g_parts = []
        g_parts = [p.strip() for p in raw_g.replace(";", ",").split(",") if p.strip()] #input in Liste einteilen (als string) und leere Einträge entfernen
        for i, g in enumerate(g_parts):
            if not g:
                continue
            gas_edges.append((int(Gas_obj.id), id_to_idx[g])) #Edges für Gasverbindung erstellen, von Gasquelle (Gas_obj.id) zu Ziel (g) aber als int

    if other:
        config.OTHER_CELL = to_internal_xy(float(other[0].x), float(other[0].y))
        Other_obj = other[0] #objekt aus liste holen, da es nur eines geben soll laut excel

        other_edges: List[Tuple[Optional[int], Optional[int]]] = [] 
        raw_o = str(Other_obj.input or "").strip()
        o_parts = []
        o_parts = [p.strip() for p in raw_o.replace(";", ",").split(",") if p.strip()] #input in Liste einteilen (als string) und leere Einträge entfernen
        for i, o in enumerate(o_parts):
            if not o:
                print(f"WARN: Other.input enthält leeren Eintrag bei Zeile {i} in Other.input")
                continue
            other_edges.append((int(Other_obj.id), id_to_idx[o])) #Edges für Otherverbindung erstellen, von Otherquelle (Other_obj.id) zu Ziel (o) aber als int

    config.WATER_CONNECTIONS = water_edges if water else []
    config.GAS_CONNECTIONS = gas_edges if gas else []
    config.OTHER_CONNECTIONS = other_edges if other else []

    #======================================================================================
    #====================Hier werden Hindernisse gesetzt===================================

    obstacles: set[Tuple[int, int]] = set()

    # 1) Säulen etc. wie bisher blockieren
    for c in cols:
        cx, cy = to_internal_xy(c.x, c.y)
        w = max(0.05, float(c.w))
        h = max(0.05, float(c.d))
        obstacles |= _cells_for_rotated_rect(cx, cy, w, h, float(c.rot))

    # 2) Alles außerhalb der durch Walls begrenzten Fläche blockieren (L-Form etc.)
    obstacles |= _outside_cells_from_walls(to_internal_xy, walls)

    config.OBSTACLES = {
        cell
        for cell in obstacles
        if 0 <= cell[0] < config.GRID_COLS and 0 <= cell[1] < config.GRID_ROWS
    }


    for i in range(len(config.MACHINE_SIZES)):
        w_m, d_m = config.MACHINE_SIZES[i]
        pd = config.MACHINE_PORTS[i] if i < len(config.MACHINE_PORTS) else {}
        port_world_xy(
            center_x=0.0,
            center_y=0.0,
            w_m=w_m,
            d_m=d_m,
            side=pd.get("side_in", "Left"),
            offset_m=pd.get("offset_in", d_m / 2.0),
            rotation_deg=0,
        )
        port_world_xy(
            center_x=0.0,
            center_y=0.0,
            w_m=w_m,
            d_m=d_m,
            side=pd.get("side_out", "Right"),
            offset_m=pd.get("offset_out", d_m / 2.0),
            rotation_deg=0,
        )

    return meta
